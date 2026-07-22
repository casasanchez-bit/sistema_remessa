"""
Gera a planilha Excel para importação em massa dos cadastros.
Execute: python gerar_planilha_importacao.py
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import sqlite3, os, re

def proximo_codigo(codigos_existentes, prefixo, digitos=4):
    """Dado um conjunto de códigos (ex: ['CE-0001','CE-0036']),
    retorna o próximo na sequência (ex: 'CE-0037').
    Se não encontrar padrão numérico, retorna prefixo + '0001'."""
    maior = 0
    for cod in codigos_existentes:
        m = re.search(r'(\d+)$', str(cod))
        if m:
            maior = max(maior, int(m.group(1)))
    return f"{prefixo}{maior + 1:0{digitos}d}"

def gerar_proximos_codigos(codigos_existentes, prefixo, quantidade, digitos=4):
    """Retorna lista com os próximos `quantidade` códigos após o maior existente."""
    maior = 0
    for cod in codigos_existentes:
        m = re.search(r'(\d+)$', str(cod))
        if m:
            maior = max(maior, int(m.group(1)))
    return [f"{prefixo}{maior + i:0{digitos}d}" for i in range(1, quantidade + 1)]

BASE = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE, "controle_remessa.db")
SAIDA = os.path.join(BASE, "importacao_cadastros.xlsx")

# ── cores ────────────────────────────────────────────────────────────────────
COR_CABEC   = "1A1A2E"   # azul escuro — cabeçalho de coluna obrigatória
COR_OPC     = "2D4059"   # azul médio  — cabeçalho coluna opcional
COR_INSTR   = "0F3460"   # azul instrução
COR_TITULO  = "E94560"   # vermelho Casa Sanchez
COR_ALT1    = "FFFFFF"
COR_ALT2    = "F0F4FF"
COR_VERDE   = "D6F0DC"
COR_AMARELO = "FFF8DC"
COR_CINZA   = "D9D9D9"
COR_ERRO    = "FFE0E0"
FONTE_BRANCA = Font(color="FFFFFF", bold=True, size=10)
FONTE_TITULO = Font(color=COR_TITULO, bold=True, size=13)
FONTE_INSTR  = Font(color="FFFFFF", size=9)

def borda_fina():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cabec(ws, col, texto, obrigatorio=True, largura=22):
    cor = COR_CABEC if obrigatorio else COR_OPC
    c = ws.cell(row=4, column=col, value=texto)
    c.fill = PatternFill("solid", fgColor=cor)
    c.font = FONTE_BRANCA
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borda_fina()
    ws.column_dimensions[get_column_letter(col)].width = largura

def titulo_aba(ws, texto, subtexto, cor=COR_TITULO):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 48
    c = ws.cell(row=1, column=1, value=texto)
    c.font = Font(color=cor, bold=True, size=14)
    c.alignment = Alignment(vertical="center")
    c2 = ws.cell(row=3, column=1, value=subtexto)
    c2.font = Font(color="444444", size=9, italic=True)
    c2.fill = PatternFill("solid", fgColor="FFF8DC")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def zebrar(ws, linha, ncols):
    cor = COR_ALT2 if linha % 2 == 0 else COR_ALT1
    for col in range(1, ncols + 1):
        c = ws.cell(row=linha, column=col)
        c.fill = PatternFill("solid", fgColor=cor)
        c.border = borda_fina()
        c.alignment = Alignment(vertical="center")

NLINHAS = 200  # linhas em branco para preenchimento

wb = openpyxl.Workbook()

# Conexão com o banco (usada em todas as abas)
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

# ══════════════════════════════════════════════════════════════════════════════
# ABA 0 — CATÁLOGO DE CORES / ESTAMPAS
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "0-Catalogo_Cores"
ws0.freeze_panes = "A5"

titulo_aba(ws0,
    "CATÁLOGO DE CORES / ESTAMPAS",
    "✅ Descrição: nome único da cor ou estampa (ex: AZUL ROYAL, FLORAL VERÃO).   "
    "⚠️ Nomes duplicados serão ignorados na importação.   "
    "⚠️ Não altere a linha 4 (cabeçalho). Preencha a partir da linha 5."
)
ws0.merge_cells("A3:A3")

cabec(ws0, 1, "DESCRIÇÃO *", largura=40)

existentes0 = cur.execute("SELECT descricao FROM catalogo_cores ORDER BY descricao").fetchall()
for i, row in enumerate(existentes0):
    ln = 5 + i
    ws0.cell(ln, 1, row["descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws0.cell(ln, 1).border = borda_fina()
    ws0.cell(ln, 1).alignment = Alignment(vertical="center")

inicio_vazio0 = 5 + len(existentes0)
for ln in range(inicio_vazio0, inicio_vazio0 + NLINHAS):
    zebrar(ws0, ln, 1)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 0b — CATÁLOGO DE SERVIÇOS
# ══════════════════════════════════════════════════════════════════════════════
ws0b = wb.create_sheet("0b-Catalogo_Servicos")
ws0b.freeze_panes = "A5"

titulo_aba(ws0b,
    "CATÁLOGO DE SERVIÇOS",
    "Descricao: nome único do serviço (ex: COSTURA E OVERLOQUE COMPLETO).   "
    "Nomes duplicados serão ignorados na importação.   "
    "Não altere a linha 4 (cabeçalho). Preencha a partir da linha 5."
)
ws0b.merge_cells("A3:A3")

cabec(ws0b, 1, "DESCRIÇÃO *", largura=50)

existentes0b = cur.execute("SELECT descricao FROM catalogo_servicos ORDER BY descricao").fetchall()
for i, row in enumerate(existentes0b):
    ln = 5 + i
    ws0b.cell(ln, 1, row["descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws0b.cell(ln, 1).border = borda_fina()
    ws0b.cell(ln, 1).alignment = Alignment(vertical="center")

inicio_vazio0b = 5 + len(existentes0b)
for ln in range(inicio_vazio0b, inicio_vazio0b + NLINHAS):
    zebrar(ws0b, ln, 1)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — PRESTADORES DE SERVIÇO (terceirizados)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1-Prestadores")
ws1.title = "1-Prestadores"
ws1.freeze_panes = "A5"

titulo_aba(ws1,
    "PRESTADORES DE SERVIÇO (Terceirizados)",
    "✅ Código: único, sem espaços (ex: C001, D001).   ✅ Nome: nome completo do prestador.   "
    "📞 Telefone: opcional, formato livre (ex: (11) 99999-9999).   "
    "⚠️ Não altere a linha 4 (cabeçalho). Preencha a partir da linha 5."
)
ws1.merge_cells("A3:C3")

cabec(ws1, 1, "CÓDIGO *", largura=14)
cabec(ws1, 2, "NOME *", largura=32)
cabec(ws1, 3, "TELEFONE", obrigatorio=False, largura=20)

existentes1 = cur.execute("SELECT codigo, nome, telefone FROM terceirizados ORDER BY codigo").fetchall()
for i, row in enumerate(existentes1):
    ln = 5 + i
    ws1.cell(ln, 1, row["codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws1.cell(ln, 2, row["nome"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws1.cell(ln, 3, row["telefone"] or "").fill = PatternFill("solid", fgColor=COR_VERDE)
    for col in range(1, 4):
        ws1.cell(ln, col).border = borda_fina()
        ws1.cell(ln, col).alignment = Alignment(vertical="center")

inicio_vazio1 = 5 + len(existentes1)
for ln in range(inicio_vazio1, inicio_vazio1 + NLINHAS):
    zebrar(ws1, ln, 3)


# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — PRODUTOS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2-Produtos")
ws2.freeze_panes = "A5"

titulo_aba(ws2,
    "PRODUTOS",
    "✅ Código: único, sem espaços (ex: 47, 48, P001).   ✅ Descrição: nome do produto.   "
    "⚠️ O Código do Produto é referenciado nas abas de Cores/Estampas e Serviços — use exatamente o mesmo código lá."
)
ws2.merge_cells("A3:B3")

cabec(ws2, 1, "CÓDIGO *", largura=14)
cabec(ws2, 2, "DESCRIÇÃO *", largura=40)

existentes2 = cur.execute("SELECT codigo, descricao FROM produtos ORDER BY codigo").fetchall()
for i, row in enumerate(existentes2):
    ln = 5 + i
    ws2.cell(ln, 1, row["codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws2.cell(ln, 2, row["descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    for col in range(1, 3):
        ws2.cell(ln, col).border = borda_fina()
        ws2.cell(ln, col).alignment = Alignment(vertical="center")

inicio_vazio2 = 5 + len(existentes2)
for ln in range(inicio_vazio2, inicio_vazio2 + NLINHAS):
    zebrar(ws2, ln, 2)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — CORES / ESTAMPAS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3-Cores_Estampas")
ws3.freeze_panes = "A5"

titulo_aba(ws3,
    "CORES / ESTAMPAS",
    "✅ Código: único (ex: C001, EST-A).   ✅ Código do Produto: deve ser idêntico ao código cadastrado na aba 2-Produtos.   "
    "✅ Descrição: nome da cor ou estampa (ex: AZUL ROYAL, FLORAL VERÃO).   "
    "⚠️ Cada cor/estampa pertence a UM produto. Se a mesma cor vale para dois produtos, adicione duas linhas com códigos diferentes."
)
ws3.merge_cells("A3:C3")

cabec(ws3, 1, "CÓDIGO *", largura=14)
cabec(ws3, 2, "CÓDIGO DO PRODUTO *", largura=22)
cabec(ws3, 3, "DESCRIÇÃO *", largura=36)

existentes3 = cur.execute("""
    SELECT ce.codigo, p.codigo AS produto_codigo, ce.descricao
    FROM cores_estampas ce JOIN produtos p ON ce.produto_id = p.id
    ORDER BY p.codigo, ce.codigo
""").fetchall()
for i, row in enumerate(existentes3):
    ln = 5 + i
    ws3.cell(ln, 1, row["codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws3.cell(ln, 2, row["produto_codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws3.cell(ln, 3, row["descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    for col in range(1, 4):
        ws3.cell(ln, col).border = borda_fina()
        ws3.cell(ln, col).alignment = Alignment(vertical="center")

codigos3 = [r["codigo"] for r in existentes3]
proximos3 = gerar_proximos_codigos(codigos3, "CE-", NLINHAS)
inicio_vazio3 = 5 + len(existentes3)
for i, ln in enumerate(range(inicio_vazio3, inicio_vazio3 + NLINHAS)):
    zebrar(ws3, ln, 3)
    c = ws3.cell(ln, 1, proximos3[i])
    c.font = Font(color="AAAAAA", italic=True, size=9)  # cinza claro = sugestão

# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — SERVIÇOS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4-Servicos")
ws4.freeze_panes = "A5"

titulo_aba(ws4,
    "SERVIÇOS",
    "✅ Código: único (ex: S001, COST-47).   ✅ Código do Produto: mesmo código da aba 2-Produtos.   "
    "✅ Descrição: nome do serviço (ex: COSTURA, BORDADO).   "
    "✅ Valor COM Registro (R$): preço por peça com registro — use ponto decimal (ex: 2.50).   "
    "⚪ Valor SEM Registro (R$): opcional — se vazio, usa o mesmo valor COM registro.   "
    "⚠️ Cada serviço pertence a UM produto. Um produto pode ter múltiplos serviços."
)
ws4.merge_cells("A3:E3")

cabec(ws4, 1, "CÓDIGO *", largura=14)
cabec(ws4, 2, "CÓDIGO DO PRODUTO *", largura=22)
cabec(ws4, 3, "DESCRIÇÃO *", largura=32)
cabec(ws4, 4, "VALOR COM REGISTRO R$ *", largura=22)
cabec(ws4, 5, "VALOR SEM REGISTRO R$", obrigatorio=False, largura=22)

existentes4 = cur.execute("""
    SELECT s.codigo, p.codigo AS produto_codigo, s.descricao,
           s.valor_com_registro, s.valor_sem_registro
    FROM servicos s JOIN produtos p ON s.produto_id = p.id
    ORDER BY p.codigo, s.codigo
""").fetchall()
for i, row in enumerate(existentes4):
    ln = 5 + i
    ws4.cell(ln, 1, row["codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws4.cell(ln, 2, row["produto_codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws4.cell(ln, 3, row["descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    c = ws4.cell(ln, 4, row["valor_com_registro"])
    c.fill = PatternFill("solid", fgColor=COR_VERDE)
    c.number_format = '#,##0.00'
    c2 = ws4.cell(ln, 5, row["valor_sem_registro"])
    c2.fill = PatternFill("solid", fgColor=COR_VERDE)
    c2.number_format = '#,##0.00'
    for col in range(1, 6):
        ws4.cell(ln, col).border = borda_fina()
        ws4.cell(ln, col).alignment = Alignment(vertical="center")

codigos4 = [r["codigo"] for r in existentes4]
proximos4 = gerar_proximos_codigos(codigos4, "S-", NLINHAS)
inicio_vazio4 = 5 + len(existentes4)
for i, ln in enumerate(range(inicio_vazio4, inicio_vazio4 + NLINHAS)):
    zebrar(ws4, ln, 5)
    c = ws4.cell(ln, 1, proximos4[i])
    c.font = Font(color="AAAAAA", italic=True, size=9)
    ws4.cell(ln, 4).number_format = '#,##0.00'
    ws4.cell(ln, 5).number_format = '#,##0.00'

# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — COMPOSIÇÃO DE PRODUTOS (Ficha Técnica de MP)
# ══════════════════════════════════════════════════════════════════════════════
ws5mp = wb.create_sheet("5-Composicao_MP")
ws5mp.freeze_panes = "A5"

titulo_aba(ws5mp,
    "COMPOSIÇÃO DE PRODUTOS — Ficha Técnica de Matéria-Prima",
    "✅ Código do Produto: mesmo código da aba 2-Produtos (ex: 318).   "
    "✅ Código MP: código da matéria-prima no ERP (ex: 10237).   "
    "✅ Descrição MP: nome da matéria-prima.   "
    "✅ Quantidade: por unidade do produto — use ponto decimal (ex: 0.33).   "
    "⚠️ Não altere a linha 4 (cabeçalho). Preencha a partir da linha 5."
)
ws5mp.merge_cells("A3:D3")

cabec(ws5mp, 1, "CÓDIGO DO PRODUTO *", largura=22)
cabec(ws5mp, 2, "CÓDIGO MP *", largura=16)
cabec(ws5mp, 3, "DESCRIÇÃO MP *", largura=46)
cabec(ws5mp, 4, "QUANTIDADE *", largura=16)

existentes5 = cur.execute("""
    SELECT p.codigo AS produto_codigo, mp.codigo AS mp_codigo,
           mp.descricao AS mp_descricao, pc.quantidade
    FROM produto_composicao pc
    JOIN produtos p ON pc.produto_id = p.id
    JOIN materias_primas mp ON pc.materia_prima_id = mp.id
    ORDER BY p.codigo, mp.codigo
""").fetchall()
for i, row in enumerate(existentes5):
    ln = 5 + i
    ws5mp.cell(ln, 1, row["produto_codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws5mp.cell(ln, 2, row["mp_codigo"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    ws5mp.cell(ln, 3, row["mp_descricao"]).fill = PatternFill("solid", fgColor=COR_VERDE)
    c = ws5mp.cell(ln, 4, row["quantidade"])
    c.fill = PatternFill("solid", fgColor=COR_VERDE)
    c.number_format = '#,##0.000'
    for col in range(1, 5):
        ws5mp.cell(ln, col).border = borda_fina()
        ws5mp.cell(ln, col).alignment = Alignment(vertical="center")

inicio_vazio5 = 5 + len(existentes5)
for ln in range(inicio_vazio5, inicio_vazio5 + NLINHAS):
    zebrar(ws5mp, ln, 4)
    ws5mp.cell(ln, 4).number_format = '#,##0.000'

conn.close()

# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — INSTRUÇÕES GERAIS
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("LEIA-ME")  # noqa — era aba 5, agora é 6 mas mantemos o nome
ws5.sheet_properties.tabColor = "E94560"
ws5.column_dimensions["A"].width = 90
ws5.row_dimensions[1].height = 36

linhas = [
    (1,  "📋  INSTRUÇÕES DE PREENCHIMENTO — Casa Sanchez · Importação em Massa", True, 14, COR_TITULO),
    (2,  "", False, 10, None),
    (3,  "ORDEM DE PREENCHIMENTO", True, 11, "1A1A2E"),
    (4,  "  1️⃣  Comece pela aba  2-Produtos  — todos os outros dependem do código do produto.", False, 10, None),
    (5,  "  2️⃣  Depois preencha  3-Cores_Estampas  e  4-Servicos  (cada linha precisa do Código do Produto já cadastrado).", False, 10, None),
    (6,  "  3️⃣  Por último, preencha  1-Prestadores  (não depende de nada).", False, 10, None),
    (7,  "", False, 10, None),
    (8,  "REGRAS GERAIS", True, 11, "1A1A2E"),
    (9,  "  • Campos com * são OBRIGATÓRIOS — deixar em branco vai gerar erro na importação.", False, 10, None),
    (10, "  • Códigos devem ser únicos dentro de cada aba — não repita o mesmo código.", False, 10, None),
    (11, "  • Não altere, não exclua e não mova a linha 4 (cabeçalhos azuis).", False, 10, None),
    (12, "  • Não insira linhas acima da linha 5.", False, 10, None),
    (13, "  • Linhas em VERDE já existem no banco — você pode incluir novas abaixo delas.", False, 10, None),
    (14, "  • Não precisa reenviar os dados verdes — o sistema ignora cadastros duplicados automaticamente.", False, 10, None),
    (15, "", False, 10, None),
    (16, "FORMATO DOS CAMPOS", True, 11, "1A1A2E"),
    (17, "  • Código: letras maiúsculas e números, sem espaços (ex: C001, 47, DAIANE).", False, 10, None),
    (18, "  • Valor (Serviços): número com ponto decimal — ex: 2.50  ou  10.00  (NÃO use vírgula).", False, 10, None),
    (19, "  • Telefone: formato livre — (11) 99999-9999 ou 11999999999.", False, 10, None),
    (20, "", False, 10, None),
    (21, "RELAÇÃO ENTRE AS TABELAS", True, 11, "1A1A2E"),
    (22, "  Produto 47 ─┬─ Cor: AZUL (C001)    ← aba 3", False, 10, None),
    (23, "              ├─ Cor: VERMELHO (C002) ← aba 3", False, 10, None),
    (24, "              ├─ Serviço: COSTURA (S001) R$ 2,50  ← aba 4", False, 10, None),
    (25, "              └─ Serviço: BORDADO (S002) R$ 5,00  ← aba 4", False, 10, None),
    (26, "", False, 10, None),
    (27, "  Produto 48 ─┬─ Cor: FLORAL (C003)  ← aba 3", False, 10, None),
    (28, "              └─ Serviço: COSTURA (S003) R$ 3,00  ← aba 4", False, 10, None),
    (29, "", False, 10, None),
    (30, "APÓS PREENCHER", True, 11, "1A1A2E"),
    (31, "  Salve o arquivo e envie para o Claude — ele importará todos os dados automaticamente.", False, 10, None),
    (32, "  Se houver erros (código duplicado, produto não encontrado), o Claude vai informar linha a linha.", False, 10, None),
]

for ln, texto, bold, size, cor in linhas:
    c = ws5.cell(row=ln, column=1, value=texto)
    c.font = Font(bold=bold, size=size, color=cor or "222222")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws5.row_dimensions[ln].height = 18 if not bold else 22
    if bold and cor == "1A1A2E":
        c.fill = PatternFill("solid", fgColor="E8EEFF")

wb.save(SAIDA)
print(f"OK Planilha gerada: {SAIDA}")
