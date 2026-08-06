import functools
import os
import shutil
import sqlite3
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, flash, g, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "controle_remessa.db"
SCHEMA_PATH = BASE_DIR / "schema.sql"
BACKUP_DIR = BASE_DIR / "backups"
FOTOS_DIR = BASE_DIR / "static" / "fotos_estampas"
FOTOS_DIR.mkdir(parents=True, exist_ok=True)
FOTOS_PLANO_DIR = BASE_DIR / "static" / "fotos_plano_corte"
FOTOS_PLANO_DIR.mkdir(parents=True, exist_ok=True)

EXTENSOES_FOTO = {".jpg", ".jpeg", ".png", ".webp"}

PERMISSOES_DISPONIVEIS = [
    ("ver_fechamento", "Ver menu Fechamento Mensal"),
    ("ver_exportar_importar", "Ver menu Exportar/Importar em Massa"),
    ("ver_historico", "Ver menu Histórico"),
    ("ver_ajuda", "Ver menu Ajuda"),
    ("ver_aprender_sql", "Ver menu Aprender SQL"),
    ("alterar_excluir_remessa", "Editar/Excluir Remessas e itens"),
    ("alterar_excluir_retorno", "Editar/Excluir Retornos"),
    ("alterar_excluir_fechamento", "Pagar/Desfazer Fechamento"),
    ("gerenciar_usuarios", "Gerenciar Usuários do Sistema (ver, criar, senhas e permissões)"),
]


def _migrar_banco():
    con = sqlite3.connect(DB_PATH)
    migrações = [
        "ALTER TABLE cores_estampas ADD COLUMN foto TEXT",
        """CREATE TABLE IF NOT EXISTS materias_primas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE,
            descricao TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS produto_composicao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id INTEGER NOT NULL REFERENCES produtos(id),
            materia_prima_id INTEGER NOT NULL REFERENCES materias_primas(id),
            quantidade REAL NOT NULL,
            UNIQUE(produto_id, materia_prima_id)
        )""",
        """CREATE TABLE IF NOT EXISTS planos_corte (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id INTEGER NOT NULL UNIQUE REFERENCES produtos(id),
            tipo_tecido TEXT,
            fornecedores TEXT,
            largura_tecido REAL,
            comprimento_enfesto REAL,
            num_camadas INTEGER,
            pecas_por_enfesto INTEGER,
            aproveitamento REAL,
            observacoes TEXT,
            foto_produto TEXT,
            foto_encaixe TEXT
        )""",
        "ALTER TABLE planos_corte ADD COLUMN foto_produto TEXT",
        "ALTER TABLE planos_corte ADD COLUMN foto_encaixe TEXT",
        "ALTER TABLE planos_corte ADD COLUMN largura_corte_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN comprimento_corte_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN tecido_dobrado INTEGER DEFAULT 0",
        "ALTER TABLE planos_corte ADD COLUMN marca TEXT",
        "ALTER TABLE planos_corte ADD COLUMN colecao TEXT",
        "ALTER TABLE planos_corte ADD COLUMN tamanho_final TEXT",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor1 TEXT",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor2 TEXT",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor3 TEXT",
        "ALTER TABLE planos_corte ADD COLUMN tipo_ziper TEXT",
        "ALTER TABLE planos_corte ADD COLUMN puxadas_min INTEGER",
        "ALTER TABLE planos_corte ADD COLUMN puxadas_max INTEGER",
        "ALTER TABLE planos_corte ADD COLUMN formato_produto TEXT DEFAULT 'retangular'",
        "ALTER TABLE planos_corte ADD COLUMN diametro_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN qtd_redondas_por_retangular INTEGER DEFAULT 1",
        "ALTER TABLE materias_primas ADD COLUMN unidade TEXT NOT NULL DEFAULT 'UN'",
        "ALTER TABLE planos_corte ADD COLUMN tem_frente_fundo INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE planos_corte ADD COLUMN largura_frente_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN comprimento_frente_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN largura_fundo_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN comprimento_fundo_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN largura_fundo_menor_cm REAL",
        "ALTER TABLE planos_corte ADD COLUMN comprimento_fundo_menor_cm REAL",
        "ALTER TABLE remessas ADD COLUMN usuario_nome TEXT",
        "ALTER TABLE retornos ADD COLUMN usuario_nome TEXT",
        """CREATE TABLE IF NOT EXISTS permissoes_usuario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL REFERENCES usuarios(id),
            permissao TEXT NOT NULL,
            UNIQUE(usuario_id, permissao)
        )""",
        """CREATE TABLE IF NOT EXISTS fornecedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE,
            nome_razao_social TEXT NOT NULL,
            telefone_empresa TEXT,
            nome_representante TEXT,
            telefone_representante TEXT
        )""",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor1_id INTEGER REFERENCES fornecedores(id)",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor2_id INTEGER REFERENCES fornecedores(id)",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor3_id INTEGER REFERENCES fornecedores(id)",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor4_id INTEGER REFERENCES fornecedores(id)",
        "ALTER TABLE planos_corte ADD COLUMN fornecedor5_id INTEGER REFERENCES fornecedores(id)",
        "ALTER TABLE itens_remessa ADD COLUMN previsao_entrega DATE",
    ]
    for sql in migrações:
        try:
            con.execute(sql)
            con.commit()
        except sqlite3.OperationalError:
            pass

    # ── Multi-serviço por item de remessa ──────────────────────────────────────
    try:
        con.execute("""CREATE TABLE IF NOT EXISTS item_servicos_remessa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_remessa_id INTEGER NOT NULL REFERENCES itens_remessa(id),
            servico_id INTEGER NOT NULL REFERENCES servicos(id),
            UNIQUE(item_remessa_id, servico_id)
        )""")
    except Exception:
        pass
    try:
        cols_ir = [r[1] for r in con.execute("PRAGMA table_info(itens_remessa)").fetchall()]
        if 'servico_id' in cols_ir:
            # Migrar servicos existentes
            con.execute("""INSERT OR IGNORE INTO item_servicos_remessa (item_remessa_id, servico_id)
                            SELECT id, servico_id FROM itens_remessa WHERE servico_id IS NOT NULL""")
            # Recriar itens_remessa sem servico_id
            con.execute("""CREATE TABLE itens_remessa_v2 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                remessa_id INTEGER NOT NULL REFERENCES remessas(id),
                produto_id INTEGER NOT NULL REFERENCES produtos(id),
                cor_estampa_id INTEGER NOT NULL REFERENCES cores_estampas(id),
                qtd_enviada INTEGER NOT NULL,
                finalizada INTEGER NOT NULL DEFAULT 0,
                prioridade INTEGER
            )""")
            con.execute("""INSERT INTO itens_remessa_v2
                            SELECT id, remessa_id, produto_id, cor_estampa_id, qtd_enviada, finalizada, prioridade
                            FROM itens_remessa""")
            con.execute("DROP TABLE itens_remessa")
            con.execute("ALTER TABLE itens_remessa_v2 RENAME TO itens_remessa")
    except Exception:
        pass
    # ── Cronometragem de corte ────────────────────────────────────────────────
    try:
        con.execute("""CREATE TABLE IF NOT EXISTS cronometragem_corte (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id    INTEGER NOT NULL REFERENCES produtos(id) ON DELETE CASCADE,
            numero        INTEGER NOT NULL CHECK(numero IN (1,2,3)),
            tempo_horas   INTEGER DEFAULT 0,
            tempo_minutos INTEGER DEFAULT 0,
            qtd_cortada   INTEGER,
            num_pessoas   INTEGER,
            data_medicao  TEXT,
            nomes_pessoas TEXT,
            UNIQUE(produto_id, numero)
        )""")
    except Exception:
        pass
    for col, definition in [
        ("data_medicao",  "TEXT"),
        ("nomes_pessoas", "TEXT"),
        ("tempo_horas",   "INTEGER DEFAULT 0"),
        ("tempo_minutos", "INTEGER DEFAULT 0"),
        ("hora_inicio",   "TEXT"),
        ("hora_fim",      "TEXT"),
    ]:
        try:
            con.execute(f"ALTER TABLE cronometragem_corte ADD COLUMN {col} {definition}")
        except Exception:
            pass

    # ── Tabela de histórico de alterações ─────────────────────────────────────
    try:
        con.execute("""CREATE TABLE IF NOT EXISTS historico (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            tabela       TEXT NOT NULL,
            registro_id  INTEGER,
            descricao    TEXT NOT NULL,
            usuario_nome TEXT,
            data_hora    TEXT NOT NULL DEFAULT (datetime('now','localtime'))
        )""")
    except Exception:
        pass

    # ── Multi-lotes de pagamento por mês ──────────────────────────────────────
    # Remove UNIQUE(terceirizado_id, mes) para permitir múltiplos lotes por mês.
    # Também recria pagamentos_fechamento_retornos: o SQLite 3.26+ atualiza a FK
    # automaticamente ao renomear a tabela pai, então precisamos recriar a filha
    # para que a referência aponte para a nova pagamentos_fechamento.
    try:
        tbl_sql = con.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='pagamentos_fechamento'"
        ).fetchone()
        if tbl_sql and "UNIQUE(terceirizado_id" in tbl_sql[0]:
            con.execute("PRAGMA foreign_keys = OFF")
            con.execute("ALTER TABLE pagamentos_fechamento RENAME TO pagamentos_fechamento_old")
            con.execute("""CREATE TABLE pagamentos_fechamento (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                terceirizado_id INTEGER NOT NULL REFERENCES terceirizados(id),
                mes TEXT NOT NULL,
                data_pagamento TEXT NOT NULL
            )""")
            con.execute("INSERT INTO pagamentos_fechamento SELECT * FROM pagamentos_fechamento_old")
            # Recria a tabela filha para resetar a FK (que o SQLite atualizou para _old)
            con.execute("ALTER TABLE pagamentos_fechamento_retornos RENAME TO pagamentos_fechamento_retornos_old")
            con.execute("""CREATE TABLE pagamentos_fechamento_retornos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pagamento_id INTEGER NOT NULL REFERENCES pagamentos_fechamento(id),
                retorno_id INTEGER NOT NULL REFERENCES retornos(id),
                UNIQUE(pagamento_id, retorno_id)
            )""")
            con.execute("INSERT INTO pagamentos_fechamento_retornos SELECT * FROM pagamentos_fechamento_retornos_old")
            con.execute("DROP TABLE pagamentos_fechamento_retornos_old")
            con.execute("DROP TABLE pagamentos_fechamento_old")
            con.execute("PRAGMA foreign_keys = ON")
            con.commit()
    except Exception:
        pass
    # ── Corrigir FK de pagamentos_fechamento_retornos (caso a migration acima
    #    já tenha rodado sem recriar a tabela filha, deixando a FK apontando para
    #    pagamentos_fechamento_old que não existe mais) ──────────────────────────
    try:
        pfr_sql = con.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='pagamentos_fechamento_retornos'"
        ).fetchone()
        if pfr_sql and "pagamentos_fechamento_old" in pfr_sql[0]:
            con.execute("PRAGMA foreign_keys = OFF")
            con.execute("ALTER TABLE pagamentos_fechamento_retornos RENAME TO pagamentos_fechamento_retornos_old")
            con.execute("""CREATE TABLE pagamentos_fechamento_retornos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pagamento_id INTEGER NOT NULL REFERENCES pagamentos_fechamento(id),
                retorno_id INTEGER NOT NULL REFERENCES retornos(id),
                UNIQUE(pagamento_id, retorno_id)
            )""")
            con.execute("INSERT INTO pagamentos_fechamento_retornos SELECT * FROM pagamentos_fechamento_retornos_old")
            con.execute("DROP TABLE pagamentos_fechamento_retornos_old")
            con.execute("PRAGMA foreign_keys = ON")
            con.commit()
    except Exception:
        pass

    con.commit()
    con.close()


_migrar_banco()


def _backup_automatico():
    """Cria backup automático na inicialização se o último tiver mais de 23h."""
    BACKUP_DIR.mkdir(exist_ok=True)
    arquivos = sorted(BACKUP_DIR.glob("*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    if arquivos:
        horas = (datetime.now() - datetime.fromtimestamp(arquivos[0].stat().st_mtime)).total_seconds() / 3600
        if horas < 23:
            return
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(DB_PATH, BACKUP_DIR / f"controle_remessa_{timestamp}.db")


_backup_automatico()

ENDPOINTS_PUBLICOS = {"login", "logout", "primeiro_acesso", "static"}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "casa-sanchez-dev-only")
csrf = CSRFProtect(app)

@app.template_filter('qty')
def qty_filter(value):
    """Formata quantidade sem zeros desnecessários: 1.000→1, 1.500→1.5, 0.330→0.33"""
    if value is None:
        return ''
    s = f"{float(value):.3f}".rstrip('0').rstrip('.')
    return s


def brdate(valor):
    """Converte uma data no padrão ISO (AAAA-MM-DD ou AAAA-MM) para o padrão Brasil."""
    if not valor:
        return valor
    partes = str(valor)[:10].split("-")
    if len(partes) == 3:
        ano, mes, dia = partes
        return f"{dia}/{mes}/{ano}"
    if len(partes) == 2:
        ano, mes = partes
        return f"{mes}/{ano}"
    return valor


app.jinja_env.filters["brdate"] = brdate


def brl(valor):
    """Formata número como moeda brasileira: 1300.5 → 1.300,50"""
    if valor is None:
        return "0,00"
    try:
        s = f"{float(valor):,.2f}"          # "1,300.50"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")  # "1.300,50"
    except (ValueError, TypeError):
        return "0,00"


app.jinja_env.filters["brl"] = brl

def smart_num(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v

app.jinja_env.filters["smart_num"] = smart_num


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def registrar_historico(db, tabela, registro_id, descricao):
    usuario_nome = session.get("usuario_nome", "Sistema")
    db.execute(
        "INSERT INTO historico (tabela, registro_id, descricao, usuario_nome) VALUES (?, ?, ?, ?)",
        (tabela, registro_id, descricao, usuario_nome),
    )


def tem_permissao(chave):
    usuario_id = session.get("usuario_id")
    if not usuario_id:
        return False
    if "permissoes" not in g:
        db = get_db()
        rows = db.execute(
            "SELECT permissao FROM permissoes_usuario WHERE usuario_id = ?", (usuario_id,)
        ).fetchall()
        g.permissoes = {r["permissao"] for r in rows}
    if chave in g.permissoes:
        return True
    if chave == "gerenciar_usuarios":
        # Ninguém foi configurado com essa permissão ainda: libera pra não travar
        # o acesso à própria tela que concede as permissões.
        db = get_db()
        alguem_configurado = db.execute(
            "SELECT 1 FROM permissoes_usuario WHERE permissao = ? LIMIT 1", (chave,)
        ).fetchone()
        if alguem_configurado is None:
            return True
    return False


@app.context_processor
def injetar_permissoes():
    return {"tem_permissao": tem_permissao}


@app.before_request
def exigir_login():
    if request.endpoint in ENDPOINTS_PUBLICOS or request.endpoint is None:
        return None
    db = get_db()
    existe_usuario = db.execute("SELECT 1 FROM usuarios LIMIT 1").fetchone() is not None
    if not existe_usuario:
        return redirect(url_for("primeiro_acesso"))
    if "usuario_id" not in session:
        return redirect(url_for("login", proximo=request.path))
    return None


# ---------------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------------

@app.route("/primeiro-acesso", methods=["GET", "POST"])
def primeiro_acesso():
    db = get_db()
    if db.execute("SELECT 1 FROM usuarios LIMIT 1").fetchone() is not None:
        return redirect(url_for("login"))
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "")
        confirmar_senha = request.form.get("confirmar_senha", "")
        if not usuario or not senha:
            flash("Informe usuário e senha.", "erro")
        elif senha != confirmar_senha:
            flash("As senhas não coincidem.", "erro")
        elif len(senha) < 4:
            flash("A senha deve ter pelo menos 4 caracteres.", "erro")
        else:
            cur = db.execute(
                "INSERT INTO usuarios (usuario, senha_hash, criado_em) VALUES (?, ?, ?)",
                (usuario, generate_password_hash(senha), date.today().isoformat()),
            )
            usuario_id = cur.lastrowid
            for chave, _ in PERMISSOES_DISPONIVEIS:
                db.execute(
                    "INSERT OR IGNORE INTO permissoes_usuario (usuario_id, permissao) VALUES (?, ?)",
                    (usuario_id, chave),
                )
            db.commit()
            flash("Usuário administrador criado com sucesso. Faça login para continuar.", "sucesso")
            return redirect(url_for("login"))
    return render_template("primeiro_acesso.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    db = get_db()
    if db.execute("SELECT 1 FROM usuarios LIMIT 1").fetchone() is None:
        return redirect(url_for("primeiro_acesso"))
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "")
        row = db.execute("SELECT * FROM usuarios WHERE usuario = ?", (usuario,)).fetchone()
        if row is None or not check_password_hash(row["senha_hash"], senha):
            flash("Usuário ou senha inválidos.", "erro")
        else:
            session.clear()
            session["usuario_id"] = row["id"]
            session["usuario_nome"] = row["usuario"]
            proximo = request.form.get("proximo", "")
            destino = proximo if proximo.startswith("/") else url_for("dashboard")
            return redirect(destino)
    return render_template("login.html", proximo=request.args.get("proximo", ""))


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Usuários do sistema
# ---------------------------------------------------------------------------

@app.route("/cadastros/usuarios")
def cadastros_usuarios():
    if not tem_permissao("gerenciar_usuarios"):
        flash("Você não tem permissão para acessar esta tela.", "erro")
        return redirect(url_for("dashboard"))
    db = get_db()
    usuarios = db.execute("SELECT id, usuario, criado_em FROM usuarios ORDER BY usuario").fetchall()
    permissoes_por_usuario = {}
    for u in usuarios:
        rows = db.execute(
            "SELECT permissao FROM permissoes_usuario WHERE usuario_id = ?", (u["id"],)
        ).fetchall()
        permissoes_por_usuario[u["id"]] = {r["permissao"] for r in rows}
    return render_template(
        "cadastro_usuarios.html", usuarios=usuarios,
        permissoes_disponiveis=PERMISSOES_DISPONIVEIS,
        permissoes_por_usuario=permissoes_por_usuario,
    )


@app.route("/cadastros/usuario", methods=["POST"])
def add_usuario():
    if not tem_permissao("gerenciar_usuarios"):
        flash("Você não tem permissão para esta ação.", "erro")
        return redirect(url_for("dashboard"))
    usuario = request.form.get("usuario", "").strip()
    senha = request.form.get("senha", "")
    if not usuario or not senha:
        flash("Informe usuário e senha.", "erro")
    elif len(senha) < 4:
        flash("A senha deve ter pelo menos 4 caracteres.", "erro")
    else:
        db = get_db()
        try:
            cur = db.execute(
                "INSERT INTO usuarios (usuario, senha_hash, criado_em) VALUES (?, ?, ?)",
                (usuario, generate_password_hash(senha), date.today().isoformat()),
            )
            usuario_id = cur.lastrowid
            permissoes_marcadas = request.form.getlist("permissao")
            for chave, _ in PERMISSOES_DISPONIVEIS:
                if chave in permissoes_marcadas:
                    db.execute(
                        "INSERT OR IGNORE INTO permissoes_usuario (usuario_id, permissao) VALUES (?, ?)",
                        (usuario_id, chave),
                    )
            db.commit()
        except sqlite3.IntegrityError:
            flash(f'Já existe um usuário com o nome "{usuario}".', "erro")
    return redirect(url_for("cadastros_usuarios"))


@app.route("/usuarios/<int:usuario_id>/permissoes", methods=["POST"])
def salvar_permissoes_usuario(usuario_id):
    if not tem_permissao("gerenciar_usuarios"):
        flash("Você não tem permissão para esta ação.", "erro")
        return redirect(url_for("dashboard"))
    db = get_db()
    usuario = db.execute("SELECT * FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
    if usuario is None:
        flash("Usuário não encontrado.", "erro")
        return redirect(url_for("cadastros_usuarios"))
    permissoes_marcadas = set(request.form.getlist("permissao"))
    db.execute("DELETE FROM permissoes_usuario WHERE usuario_id = ?", (usuario_id,))
    for chave, _ in PERMISSOES_DISPONIVEIS:
        if chave in permissoes_marcadas:
            db.execute(
                "INSERT INTO permissoes_usuario (usuario_id, permissao) VALUES (?, ?)",
                (usuario_id, chave),
            )
    db.commit()
    flash(f"Permissões de {usuario['usuario']} atualizadas.", "sucesso")
    return redirect(url_for("cadastros_usuarios"))


@app.route("/usuarios/<int:usuario_id>/senha", methods=["POST"])
def alterar_senha_usuario(usuario_id):
    if not tem_permissao("gerenciar_usuarios"):
        flash("Você não tem permissão para esta ação.", "erro")
        return redirect(url_for("dashboard"))
    db = get_db()
    usuario = db.execute("SELECT * FROM usuarios WHERE id = ?", (usuario_id,)).fetchone()
    if usuario is None:
        flash("Usuário não encontrado.", "erro")
        return redirect(url_for("cadastros_usuarios"))
    nova_senha = request.form.get("nova_senha", "")
    if len(nova_senha) < 4:
        flash("A senha deve ter pelo menos 4 caracteres.", "erro")
    else:
        db.execute(
            "UPDATE usuarios SET senha_hash = ? WHERE id = ?",
            (generate_password_hash(nova_senha), usuario_id),
        )
        db.commit()
        flash(f"Senha de {usuario['usuario']} alterada com sucesso.", "sucesso")
    return redirect(url_for("cadastros_usuarios"))


@app.route("/usuarios/<int:usuario_id>/excluir", methods=["POST"])
def excluir_usuario(usuario_id):
    if not tem_permissao("gerenciar_usuarios"):
        flash("Você não tem permissão para esta ação.", "erro")
        return redirect(url_for("dashboard"))
    db = get_db()
    total = db.execute("SELECT COUNT(*) AS n FROM usuarios").fetchone()["n"]
    if total <= 1:
        flash("Não é possível excluir: precisa existir ao menos um usuário no sistema.", "erro")
    elif usuario_id == session.get("usuario_id"):
        flash("Você não pode excluir o próprio usuário enquanto estiver logado com ele.", "erro")
    else:
        db.execute("DELETE FROM permissoes_usuario WHERE usuario_id = ?", (usuario_id,))
        db.execute("DELETE FROM usuarios WHERE id = ?", (usuario_id,))
        db.commit()
    return redirect(url_for("cadastros_usuarios"))


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    try:
        conn.execute("ALTER TABLE itens_remessa ADD COLUMN prioridade INTEGER")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE cores_estampas ADD COLUMN produto_id INTEGER REFERENCES produtos(id)")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS catalogo_cores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descricao TEXT NOT NULL UNIQUE
        )""")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE servicos ADD COLUMN valor_sem_registro REAL NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE servicos RENAME COLUMN valor TO valor_com_registro")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE terceirizados ADD COLUMN registrado INTEGER NOT NULL DEFAULT 1")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS catalogo_servicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descricao TEXT NOT NULL UNIQUE
        )""")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE remessas ADD COLUMN observacao TEXT NOT NULL DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE retornos ADD COLUMN observacao TEXT NOT NULL DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


def qtd_retornada(db, item_remessa_id):
    """Soma o retornado deste item de remessa específico (Produto+Cor/Estampa+Serviço)."""
    row = db.execute(
        "SELECT COALESCE(SUM(qtd_retornada), 0) AS total FROM itens_retorno WHERE item_remessa_id = ?",
        (item_remessa_id,),
    ).fetchone()
    return row["total"]


def proximo_numero(db, tabela, coluna="numero"):
    row = db.execute(f"SELECT COALESCE(MAX({coluna}), 0) + 1 AS proximo FROM {tabela}").fetchone()
    return row["proximo"]


def proximo_codigo(db, tabela, prefixo):
    rows = db.execute(f"SELECT codigo FROM {tabela}").fetchall()
    maior = 0
    for r in rows:
        codigo = r["codigo"] or ""
        if codigo.startswith(prefixo + "-"):
            try:
                maior = max(maior, int(codigo.split("-")[1]))
            except (IndexError, ValueError):
                pass
    return f"{prefixo}-{maior + 1:04d}"


def mes_pago(db, terceirizado_id, mes):
    row = db.execute(
        "SELECT 1 FROM pagamentos_fechamento WHERE terceirizado_id = ? AND mes = ?",
        (terceirizado_id, mes),
    ).fetchone()
    return row is not None


def retorno_pago(db, retorno_id):
    """Um retorno é considerado pago somente se foi incluído no instantâneo de algum
    pagamento já confirmado (não basta estar no mesmo mês calendário de um pagamento)."""
    row = db.execute(
        "SELECT 1 FROM pagamentos_fechamento_retornos WHERE retorno_id = ?", (retorno_id,)
    ).fetchone()
    return row is not None


def item_tem_retorno_pago(db, item_remessa_id):
    """Verifica se algum retorno já pago está vinculado a este item de remessa específico."""
    retorno_ids = [
        r["retorno_id"]
        for r in db.execute(
            "SELECT DISTINCT retorno_id FROM itens_retorno WHERE item_remessa_id = ?", (item_remessa_id,)
        ).fetchall()
    ]
    return any(retorno_pago(db, rid) for rid in retorno_ids)


def remessa_tem_retorno_pago(db, remessa_id):
    item_ids = [r["id"] for r in db.execute("SELECT id FROM itens_remessa WHERE remessa_id = ?", (remessa_id,)).fetchall()]
    return any(item_tem_retorno_pago(db, item_id) for item_id in item_ids)


def remessa_paga(db, remessa_id):
    """Uma remessa só é considerada paga quando todos os seus itens estiverem com saldo
    zerado e todos os retornos vinculados a eles já tiverem sido pagos."""
    itens = db.execute("SELECT id, qtd_enviada FROM itens_remessa WHERE remessa_id = ?", (remessa_id,)).fetchall()
    if not itens:
        return False
    for item in itens:
        if item["qtd_enviada"] - qtd_retornada(db, item["id"]) > 0:
            return False
        retorno_ids = [
            r["retorno_id"]
            for r in db.execute(
                "SELECT DISTINCT retorno_id FROM itens_retorno WHERE item_remessa_id = ?", (item["id"],)
            ).fetchall()
        ]
        if not retorno_ids or any(not retorno_pago(db, rid) for rid in retorno_ids):
            return False
    return True


def recalcular_finalizada(db, item_remessa_id):
    item = db.execute("SELECT qtd_enviada FROM itens_remessa WHERE id = ?", (item_remessa_id,)).fetchone()
    if item is None:
        return
    retornado = qtd_retornada(db, item_remessa_id)
    finalizada = 1 if retornado >= item["qtd_enviada"] else 0
    db.execute("UPDATE itens_remessa SET finalizada = ? WHERE id = ?", (finalizada, item_remessa_id))


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route("/")
def dashboard():
    db = get_db()
    filtro_data_inicio = request.args.get("filtro_data_inicio", "").strip()
    filtro_data_fim = request.args.get("filtro_data_fim", "").strip()
    hoje_date = date.today()

    query = """SELECT itens_remessa.id, itens_remessa.remessa_id, itens_remessa.produto_id, itens_remessa.qtd_enviada,
                  itens_remessa.previsao_entrega,
                  remessas.terceirizado_id AS terceirizado_id,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  COALESCE(SUM(CASE WHEN terceirizados.registrado = 1
                                    THEN servicos.valor_com_registro
                                    ELSE servicos.valor_sem_registro END), 0) AS valor_unitario
           FROM itens_remessa
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           WHERE 1=1"""
    params = []
    if filtro_data_inicio:
        query += " AND remessas.data_envio >= ?"
        params.append(filtro_data_inicio)
    if filtro_data_fim:
        query += " AND remessas.data_envio <= ?"
        params.append(filtro_data_fim)
    query += " GROUP BY itens_remessa.id"
    itens = db.execute(query, params).fetchall()

    remessas_ativas_ids = set()
    retornado_por_terceirizado = {}
    retornado_por_produto = {}
    enviado_por_terceirizado = {}
    enviado_por_produto = {}
    produto_info = {}
    hoje_str = hoje_date.isoformat()
    itens_prazo_vencido = 0
    valor_total_a_pagar = 0.0
    for i in itens:
        retornado = qtd_retornada(db, i["id"])
        pendente_item = i["qtd_enviada"] - retornado
        if pendente_item > 0:
            remessas_ativas_ids.add(i["remessa_id"])
            valor_total_a_pagar += pendente_item * (i["valor_unitario"] or 0)
            if i["previsao_entrega"] and i["previsao_entrega"] < hoje_str:
                itens_prazo_vencido += 1
        retornado_por_terceirizado[i["terceirizado_id"]] = retornado_por_terceirizado.get(i["terceirizado_id"], 0) + retornado
        retornado_por_produto[i["produto_id"]] = retornado_por_produto.get(i["produto_id"], 0) + retornado
        enviado_por_terceirizado[i["terceirizado_id"]] = enviado_por_terceirizado.get(i["terceirizado_id"], 0) + i["qtd_enviada"]
        enviado_por_produto[i["produto_id"]] = enviado_por_produto.get(i["produto_id"], 0) + i["qtd_enviada"]
        produto_info[i["produto_id"]] = (i["produto_codigo"], i["produto_descricao"])
    remessas_ativas = len(remessas_ativas_ids)

    # Alerta: remessas com itens pendentes há mais de 30 dias
    todas_remessas = db.execute(
        """SELECT remessas.id, remessas.numero, remessas.data_envio, terceirizados.nome AS terceirizado_nome
           FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           ORDER BY remessas.data_envio""",
    ).fetchall()
    remessas_alerta = []
    for r in todas_remessas:
        try:
            data_envio = date.fromisoformat(r["data_envio"])
            dias = (hoje_date - data_envio).days
        except (ValueError, TypeError):
            continue
        if dias <= 30:
            continue
        itens_r = db.execute(
            "SELECT id, qtd_enviada FROM itens_remessa WHERE remessa_id = ?", (r["id"],)
        ).fetchall()
        tem_pendente = any(qtd_retornada(db, i["id"]) < i["qtd_enviada"] for i in itens_r)
        if tem_pendente:
            remessas_alerta.append({"id": r["id"], "numero": r["numero"],
                                    "terceirizado_nome": r["terceirizado_nome"],
                                    "data_envio": r["data_envio"], "dias": dias})

    terceirizados = db.execute("SELECT * FROM terceirizados ORDER BY nome").fetchall()
    saldo_por_terceirizado = []
    for t in terceirizados:
        enviado = enviado_por_terceirizado.get(t["id"], 0)
        retornado = retornado_por_terceirizado.get(t["id"], 0)
        pendente = enviado - retornado
        if pendente <= 0:
            continue
        saldo_por_terceirizado.append(
            {"id": t["id"], "nome": t["nome"], "enviado": enviado, "retornado": retornado, "pendente": pendente}
        )
    terceirizados_com_pendencia = len(saldo_por_terceirizado)

    saldo_por_produto = []
    for produto_id, (codigo, descricao) in produto_info.items():
        enviado = enviado_por_produto.get(produto_id, 0)
        retornado = retornado_por_produto.get(produto_id, 0)
        pendente = enviado - retornado
        if pendente <= 0:
            continue
        saldo_por_produto.append(
            {"id": produto_id, "codigo": codigo, "descricao": descricao, "enviado": enviado, "retornado": retornado, "pendente": pendente}
        )
    saldo_por_produto.sort(key=lambda s: s["codigo"])

    return render_template(
        "dashboard.html",
        remessas_ativas=remessas_ativas,
        terceirizados_com_pendencia=terceirizados_com_pendencia,
        itens_prazo_vencido=itens_prazo_vencido,
        valor_total_a_pagar=valor_total_a_pagar,
        saldo_por_terceirizado=saldo_por_terceirizado,
        saldo_por_produto=saldo_por_produto,
        remessas_alerta=remessas_alerta,
        filtro_data_inicio=filtro_data_inicio,
        filtro_data_fim=filtro_data_fim,
    )


# ---------------------------------------------------------------------------
# Cadastros — Terceirizados
# ---------------------------------------------------------------------------

@app.route("/cadastros")
def cadastros():
    return redirect(url_for("cadastros_terceirizados"))


@app.route("/cadastros/terceirizados")
def cadastros_terceirizados():
    db = get_db()
    terceirizados = db.execute("SELECT * FROM terceirizados ORDER BY nome").fetchall()
    return render_template("cadastro_terceirizados.html", terceirizados=terceirizados)


@app.route("/cadastros/terceirizado", methods=["POST"])
def add_terceirizado():
    nome = request.form.get("nome", "").strip()
    telefone = request.form.get("telefone", "").strip()
    registrado = 1 if request.form.get("registrado") == "1" else 0
    if nome:
        db = get_db()
        codigo = proximo_codigo(db, "terceirizados", "T")
        try:
            db.execute(
                "INSERT INTO terceirizados (codigo, nome, telefone, registrado) VALUES (?, ?, ?, ?)",
                (codigo, nome, telefone, registrado),
            )
            db.commit()
        except sqlite3.IntegrityError:
            flash(f'Já existe um terceirizado cadastrado com o nome "{nome}".', "erro")
    return redirect(url_for("cadastros_terceirizados"))


@app.route("/terceirizados/<int:terceirizado_id>")
def ver_terceirizado(terceirizado_id):
    db = get_db()
    terceirizado = db.execute("SELECT * FROM terceirizados WHERE id = ?", (terceirizado_id,)).fetchone()
    if terceirizado is None:
        flash("Terceirizado não encontrado.", "erro")
        return redirect(url_for("cadastros_terceirizados"))
    remessas_do = db.execute(
        """SELECT itens_remessa.*, remessas.numero AS numero, remessas.data_envio AS data_envio,
                  produtos.descricao AS produto_descricao, cores_estampas.descricao AS cor_descricao
           FROM itens_remessa
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           WHERE remessas.terceirizado_id = ? ORDER BY remessas.numero DESC""",
        (terceirizado_id,),
    ).fetchall()
    remessas_view = []
    for item in remessas_do:
        retornado = qtd_retornada(db, item["id"])
        pendente = item["qtd_enviada"] - retornado
        remessas_view.append({**dict(item), "qtd_retornada": retornado, "pendente": pendente})
    return render_template("terceirizado_ver.html", terceirizado=terceirizado, remessas=remessas_view)


@app.route("/terceirizados/<int:terceirizado_id>/editar", methods=["GET", "POST"])
def editar_terceirizado(terceirizado_id):
    db = get_db()
    terceirizado = db.execute("SELECT * FROM terceirizados WHERE id = ?", (terceirizado_id,)).fetchone()
    if terceirizado is None:
        flash("Terceirizado não encontrado.", "erro")
        return redirect(url_for("cadastros_terceirizados"))
    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        telefone = request.form.get("telefone", "").strip()
        registrado = 1 if request.form.get("registrado") == "1" else 0
        if nome:
            try:
                db.execute(
                    "UPDATE terceirizados SET nome = ?, telefone = ?, registrado = ? WHERE id = ?",
                    (nome, telefone, registrado, terceirizado_id),
                )
                db.commit()
                return redirect(url_for("cadastros_terceirizados"))
            except sqlite3.IntegrityError:
                flash(f'Já existe um terceirizado cadastrado com o nome "{nome}".', "erro")
    return render_template("terceirizado_editar.html", terceirizado=terceirizado)


@app.route("/terceirizados/<int:terceirizado_id>/excluir", methods=["POST"])
def excluir_terceirizado(terceirizado_id):
    db = get_db()
    em_uso = db.execute(
        "SELECT COUNT(*) AS n FROM remessas WHERE terceirizado_id = ?", (terceirizado_id,)
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: este terceirizado já tem remessas registradas.", "erro")
    else:
        db.execute("DELETE FROM terceirizados WHERE id = ?", (terceirizado_id,))
        db.commit()
    return redirect(url_for("cadastros_terceirizados"))


# ---------------------------------------------------------------------------
# Cadastros — Produtos
# ---------------------------------------------------------------------------

@app.route("/cadastros/produtos")
def cadastros_produtos():
    db = get_db()
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    return render_template("cadastro_produtos.html", produtos=produtos)


@app.route("/cadastros/produto", methods=["POST"])
def add_produto():
    codigo = request.form.get("codigo", "").strip()
    descricao = request.form.get("descricao", "").strip()
    if codigo and descricao:
        try:
            get_db().execute(
                "INSERT INTO produtos (codigo, descricao) VALUES (?, ?)",
                (codigo, descricao),
            )
            get_db().commit()
        except sqlite3.IntegrityError:
            flash(f'Já existe um produto cadastrado com o código "{codigo}".', "erro")
    return redirect(url_for("cadastros_produtos"))


@app.route("/produtos/<int:produto_id>")
def ver_produto(produto_id):
    db = get_db()
    produto = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if produto is None:
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("cadastros_produtos"))
    servicos_do = db.execute("SELECT * FROM servicos WHERE produto_id = ?", (produto_id,)).fetchall()

    itens = db.execute(
        """SELECT itens_remessa.id, itens_remessa.qtd_enviada,
                  remessas.terceirizado_id AS terceirizado_id, terceirizados.nome AS terceirizado_nome
           FROM itens_remessa
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           WHERE itens_remessa.produto_id = ?""",
        (produto_id,),
    ).fetchall()

    enviado_por_terceirizado = {}
    retornado_por_terceirizado = {}
    nome_por_terceirizado = {}
    for i in itens:
        tid = i["terceirizado_id"]
        retornado = qtd_retornada(db, i["id"])
        enviado_por_terceirizado[tid] = enviado_por_terceirizado.get(tid, 0) + i["qtd_enviada"]
        retornado_por_terceirizado[tid] = retornado_por_terceirizado.get(tid, 0) + retornado
        nome_por_terceirizado[tid] = i["terceirizado_nome"]

    pendente_por_terceirizado = []
    for tid, nome in nome_por_terceirizado.items():
        enviado = enviado_por_terceirizado.get(tid, 0)
        retornado = retornado_por_terceirizado.get(tid, 0)
        pendente = enviado - retornado
        if pendente <= 0:
            continue
        pendente_por_terceirizado.append(
            {"id": tid, "nome": nome, "enviado": enviado, "retornado": retornado, "pendente": pendente}
        )
    pendente_por_terceirizado.sort(key=lambda x: x["nome"])

    tem_plano = db.execute(
        "SELECT 1 FROM planos_corte WHERE produto_id = ? LIMIT 1", (produto_id,)
    ).fetchone() is not None
    tem_composicao = db.execute(
        "SELECT 1 FROM produto_composicao WHERE produto_id = ? LIMIT 1", (produto_id,)
    ).fetchone() is not None

    return render_template(
        "produto_ver.html",
        produto=produto,
        servicos=servicos_do,
        pendente_por_terceirizado=pendente_por_terceirizado,
        tem_plano=tem_plano,
        tem_composicao=tem_composicao,
    )


@app.route("/produtos/<int:produto_id>/editar", methods=["GET", "POST"])
def editar_produto(produto_id):
    db = get_db()
    produto = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if produto is None:
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("cadastros_produtos"))
    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        descricao = request.form.get("descricao", "").strip()
        if codigo and descricao:
            try:
                db.execute(
                    "UPDATE produtos SET codigo = ?, descricao = ? WHERE id = ?",
                    (codigo, descricao, produto_id),
                )
                db.commit()
                return redirect(url_for("cadastros_produtos"))
            except sqlite3.IntegrityError:
                flash(f'Já existe um produto cadastrado com o código "{codigo}".', "erro")
    return render_template("produto_editar.html", produto=produto)


@app.route("/produtos/<int:produto_id>/excluir", methods=["POST"])
def excluir_produto(produto_id):
    db = get_db()
    em_uso = db.execute(
        "SELECT COUNT(*) AS n FROM itens_remessa WHERE produto_id = ?", (produto_id,)
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: este produto já tem remessas registradas.", "erro")
    else:
        db.execute("DELETE FROM servicos WHERE produto_id = ?", (produto_id,))
        db.execute("DELETE FROM produtos WHERE id = ?", (produto_id,))
        db.commit()
    return redirect(url_for("cadastros_produtos"))


# ---------------------------------------------------------------------------
# Cadastros — Cores/Estampas
# ---------------------------------------------------------------------------

@app.route("/cadastros/cores-estampas")
def cadastros_cores_estampas():
    db = get_db()
    auto_vincular_fotos(db)
    busca = request.args.get("busca", "").strip()
    buscou = "busca" in request.args

    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    catalogo_cores = db.execute(
        "SELECT * FROM catalogo_cores ORDER BY descricao COLLATE NOCASE"
    ).fetchall()

    if busca:
        like = f"%{busca}%"
        cores_estampas = db.execute(
            """SELECT cores_estampas.*, produtos.descricao AS produto_descricao
               FROM cores_estampas JOIN produtos ON produtos.id = cores_estampas.produto_id
               WHERE produtos.descricao LIKE ? OR produtos.codigo LIKE ? OR cores_estampas.descricao LIKE ?
               ORDER BY CAST(SUBSTR(cores_estampas.codigo, 3) AS INTEGER), cores_estampas.descricao COLLATE NOCASE""",
            (like, like, like)
        ).fetchall()
    else:
        cores_estampas = db.execute(
            """SELECT cores_estampas.*, produtos.descricao AS produto_descricao
               FROM cores_estampas JOIN produtos ON produtos.id = cores_estampas.produto_id
               ORDER BY CAST(SUBSTR(cores_estampas.codigo, 3) AS INTEGER), cores_estampas.descricao COLLATE NOCASE"""
        ).fetchall()

    return render_template("cadastro_cores_estampas.html", produtos=produtos,
                           cores_estampas=cores_estampas, catalogo_cores=catalogo_cores,
                           busca=busca)


@app.route("/catalogo-cores/adicionar", methods=["POST"])
def add_catalogo_cor():
    descricao = request.form.get("nova_cor", "").strip().upper()
    if descricao:
        db = get_db()
        try:
            db.execute("INSERT INTO catalogo_cores (descricao) VALUES (?)", (descricao,))
            db.commit()
        except sqlite3.IntegrityError:
            flash(f'Cor/Estampa "{descricao}" já existe no catálogo.', "erro")
    return redirect(url_for("cadastros_cores_estampas"))


@app.route("/catalogo-cores/<int:cor_id>/editar", methods=["POST"])
def editar_catalogo_cor(cor_id):
    db = get_db()
    nova_descricao = request.form.get("descricao", "").strip().upper()
    if not nova_descricao:
        flash("Descrição não pode ser vazia.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    cat = db.execute("SELECT descricao FROM catalogo_cores WHERE id = ?", (cor_id,)).fetchone()
    if not cat:
        flash("Cor/estampa não encontrada.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    descricao_antiga = cat["descricao"]
    try:
        db.execute("UPDATE catalogo_cores SET descricao = ? WHERE id = ?", (nova_descricao, cor_id))
        # Sincroniza automaticamente todas as atribuições que usavam a descrição antiga
        db.execute(
            "UPDATE cores_estampas SET descricao = ? WHERE descricao = ?",
            (nova_descricao, descricao_antiga),
        )
        db.commit()
    except sqlite3.IntegrityError:
        flash(f'Já existe uma cor/estampa com o nome "{nova_descricao}" no catálogo.', "erro")
    return redirect(url_for("cadastros_cores_estampas"))


@app.route("/catalogo-cores/<int:cor_id>/excluir", methods=["POST"])
def excluir_catalogo_cor(cor_id):
    db = get_db()
    em_uso = db.execute(
        "SELECT COUNT(*) AS n FROM cores_estampas WHERE descricao = (SELECT descricao FROM catalogo_cores WHERE id = ?)",
        (cor_id,)
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: esta cor/estampa já está atribuída a um produto.", "erro")
    else:
        db.execute("DELETE FROM catalogo_cores WHERE id = ?", (cor_id,))
        db.commit()
    return redirect(url_for("cadastros_cores_estampas"))


@app.route("/cadastros/cor-estampa", methods=["POST"])
def add_cor_estampa():
    produto_id = request.form.get("produto_id")
    catalogo_cor_id = request.form.get("catalogo_cor_id")
    if produto_id and catalogo_cor_id:
        db = get_db()
        cat = db.execute("SELECT descricao FROM catalogo_cores WHERE id = ?", (catalogo_cor_id,)).fetchone()
        if not cat:
            flash("Cor/estampa não encontrada no catálogo.", "erro")
            return redirect(url_for("cadastros_cores_estampas"))
        descricao = cat["descricao"]
        ja_existe = db.execute(
            "SELECT id FROM cores_estampas WHERE produto_id = ? AND descricao = ?",
            (produto_id, descricao),
        ).fetchone()
        if ja_existe:
            flash(f"A cor/estampa '{descricao}' já está atribuída a este produto.", "erro")
        else:
            codigo = proximo_codigo(db, "cores_estampas", "CE")
            db.execute(
                "INSERT INTO cores_estampas (codigo, produto_id, descricao) VALUES (?, ?, ?)",
                (codigo, produto_id, descricao),
            )
            db.commit()
    return redirect(url_for("cadastros_cores_estampas"))


def auto_vincular_fotos(db):
    """Escaneia static/fotos_estampas/ e vincula arquivos cujo nome bate com cores_estampas.codigo."""
    for arquivo in FOTOS_DIR.iterdir():
        if arquivo.suffix.lower() in EXTENSOES_FOTO:
            codigo = arquivo.stem.upper()
            db.execute(
                "UPDATE cores_estampas SET foto = ? WHERE UPPER(codigo) = ?",
                (arquivo.name, codigo),
            )
    db.commit()


@app.route("/cores-estampas/<int:cor_id>/foto", methods=["POST"])
def upload_foto_estampa(cor_id):
    db = get_db()
    ce = db.execute("SELECT codigo FROM cores_estampas WHERE id = ?", (cor_id,)).fetchone()
    if ce is None:
        flash("Atribuição não encontrada.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    arquivo = request.files.get("foto")
    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo de imagem.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    ext = Path(arquivo.filename).suffix.lower()
    if ext not in EXTENSOES_FOTO:
        flash("Formato não suportado. Use JPG, PNG ou WEBP.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    nome_arquivo = ce["codigo"].upper() + ext
    arquivo.save(FOTOS_DIR / nome_arquivo)
    db.execute("UPDATE cores_estampas SET foto = ? WHERE id = ?", (nome_arquivo, cor_id))
    db.commit()
    flash("Foto vinculada com sucesso.", "ok")
    return redirect(url_for("cadastros_cores_estampas"))


@app.route("/cores-estampas/<int:cor_id>/foto/remover", methods=["POST"])
def remover_foto_estampa(cor_id):
    db = get_db()
    ce = db.execute("SELECT foto FROM cores_estampas WHERE id = ?", (cor_id,)).fetchone()
    if ce and ce["foto"]:
        caminho = FOTOS_DIR / ce["foto"]
        if caminho.exists():
            caminho.unlink()
        db.execute("UPDATE cores_estampas SET foto = NULL WHERE id = ?", (cor_id,))
        db.commit()
        flash("Foto removida.", "ok")
    return redirect(url_for("cadastros_cores_estampas"))


@app.route("/cores-estampas/<int:cor_id>")
def ver_cor_estampa(cor_id):
    db = get_db()
    cor = db.execute(
        """SELECT cores_estampas.*, produtos.descricao AS produto_descricao
           FROM cores_estampas JOIN produtos ON produtos.id = cores_estampas.produto_id WHERE cores_estampas.id = ?""",
        (cor_id,),
    ).fetchone()
    if cor is None:
        flash("Cor/Estampa não encontrada.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    return render_template("cor_estampa_ver.html", cor=cor)


@app.route("/cores-estampas/<int:cor_id>/editar", methods=["GET", "POST"])
def editar_cor_estampa(cor_id):
    db = get_db()
    cor = db.execute("SELECT * FROM cores_estampas WHERE id = ?", (cor_id,)).fetchone()
    if cor is None:
        flash("Cor/Estampa não encontrada.", "erro")
        return redirect(url_for("cadastros_cores_estampas"))
    if request.method == "POST":
        descricao = request.form.get("descricao", "").strip()
        if descricao:
            db.execute("UPDATE cores_estampas SET descricao = ? WHERE id = ?", (descricao, cor_id))
            db.commit()
            return redirect(url_for("cadastros_cores_estampas"))
    return render_template("cor_estampa_editar.html", cor=cor)


@app.route("/cores-estampas/<int:cor_id>/excluir", methods=["POST"])
def excluir_cor_estampa(cor_id):
    db = get_db()
    em_uso = db.execute(
        "SELECT COUNT(*) AS n FROM itens_remessa WHERE cor_estampa_id = ?", (cor_id,)
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: esta cor/estampa já tem remessas registradas.", "erro")
    else:
        db.execute("DELETE FROM cores_estampas WHERE id = ?", (cor_id,))
        db.commit()
    return redirect(url_for("cadastros_cores_estampas"))


# ---------------------------------------------------------------------------
# Cadastros — Serviços
# ---------------------------------------------------------------------------

@app.route("/cadastros/servicos")
def cadastros_servicos():
    db = get_db()
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    catalogo_servicos = db.execute(
        "SELECT * FROM catalogo_servicos ORDER BY descricao COLLATE NOCASE"
    ).fetchall()
    servicos = db.execute(
        """SELECT servicos.*, produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao
           FROM servicos JOIN produtos ON produtos.id = servicos.produto_id
           ORDER BY produtos.descricao, servicos.descricao COLLATE NOCASE"""
    ).fetchall()
    return render_template("cadastro_servicos.html", produtos=produtos, servicos=servicos,
                           catalogo_servicos=catalogo_servicos)


@app.route("/catalogo-servicos/adicionar", methods=["POST"])
def add_catalogo_servico():
    descricao = request.form.get("novo_servico", "").strip().upper()
    if descricao:
        db = get_db()
        try:
            db.execute("INSERT INTO catalogo_servicos (descricao) VALUES (?)", (descricao,))
            db.commit()
        except sqlite3.IntegrityError:
            flash(f'"{descricao}" já existe no catálogo.', "erro")
    return redirect(url_for("cadastros_servicos"))


@app.route("/catalogo-servicos/<int:servico_id>/editar", methods=["POST"])
def editar_catalogo_servico(servico_id):
    db = get_db()
    nova_descricao = request.form.get("descricao", "").strip().upper()
    if not nova_descricao:
        flash("Descrição não pode ser vazia.", "erro")
        return redirect(url_for("cadastros_servicos"))
    cat = db.execute("SELECT descricao FROM catalogo_servicos WHERE id = ?", (servico_id,)).fetchone()
    if not cat:
        flash("Serviço não encontrado.", "erro")
        return redirect(url_for("cadastros_servicos"))
    descricao_antiga = cat["descricao"]
    try:
        db.execute("UPDATE catalogo_servicos SET descricao = ? WHERE id = ?", (nova_descricao, servico_id))
        # Sincroniza automaticamente todas as atribuições que usavam a descrição antiga
        db.execute(
            "UPDATE servicos SET descricao = ? WHERE descricao = ?",
            (nova_descricao, descricao_antiga),
        )
        db.commit()
    except sqlite3.IntegrityError:
        flash(f"Já existe um serviço com o nome '{nova_descricao}' no catálogo.", "erro")
    return redirect(url_for("cadastros_servicos"))


@app.route("/catalogo-servicos/<int:servico_id>/excluir", methods=["POST"])
def excluir_catalogo_servico(servico_id):
    db = get_db()
    db.execute("DELETE FROM catalogo_servicos WHERE id = ?", (servico_id,))
    db.commit()
    return redirect(url_for("cadastros_servicos"))


@app.route("/cadastros/servico", methods=["POST"])
def add_servico():
    produto_id = request.form.get("produto_id")
    descricao = request.form.get("descricao", "").strip().upper()
    valor_com = request.form.get("valor_com_registro", "0").strip()
    valor_sem = request.form.get("valor_sem_registro", "0").strip()
    if not produto_id or not descricao:
        flash("Produto e serviço são obrigatórios.", "erro")
        return redirect(url_for("cadastros_servicos"))
    try:
        v_com = float(valor_com or 0)
        v_sem = float(valor_sem or 0)
    except ValueError:
        v_com = v_sem = 0
    if v_com <= 0 or v_sem <= 0:
        flash("Os valores com e sem registro devem ser maiores que zero.", "erro")
        return redirect(url_for("cadastros_servicos"))
    db = get_db()
    existente = db.execute(
        "SELECT id FROM servicos WHERE produto_id = ? AND descricao = ?", (produto_id, descricao)
    ).fetchone()
    if existente:
        flash(f'O serviço "{descricao}" já está atribuído a este produto. Para alterar o valor, use VISUALIZAR → EDITAR.', "erro")
        return redirect(url_for("cadastros_servicos"))
    codigo = proximo_codigo(db, "servicos", "SV")
    db.execute(
        "INSERT INTO servicos (codigo, produto_id, descricao, valor_com_registro, valor_sem_registro) VALUES (?, ?, ?, ?, ?)",
        (codigo, produto_id, descricao, v_com, v_sem),
    )
    db.commit()
    return redirect(url_for("cadastros_servicos"))


@app.route("/servicos/<int:servico_id>")
def ver_servico(servico_id):
    db = get_db()
    servico = db.execute(
        """SELECT servicos.*, produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao
           FROM servicos JOIN produtos ON produtos.id = servicos.produto_id WHERE servicos.id = ?""",
        (servico_id,),
    ).fetchone()
    if servico is None:
        flash("Serviço não encontrado.", "erro")
        return redirect(url_for("cadastros_servicos"))
    return render_template("servico_ver.html", servico=servico)


@app.route("/servicos/<int:servico_id>/editar", methods=["GET", "POST"])
def editar_servico(servico_id):
    db = get_db()
    servico = db.execute("SELECT * FROM servicos WHERE id = ?", (servico_id,)).fetchone()
    if servico is None:
        flash("Serviço não encontrado.", "erro")
        return redirect(url_for("cadastros_servicos"))
    if request.method == "POST":
        valor_com = request.form.get("valor_com_registro", "0").strip()
        valor_sem = request.form.get("valor_sem_registro", "0").strip()
        try:
            v_com = float(valor_com or 0)
            v_sem = float(valor_sem or 0)
        except ValueError:
            v_com = v_sem = 0
        if v_com <= 0 or v_sem <= 0:
            flash("Os valores com e sem registro devem ser maiores que zero.", "erro")
            return redirect(url_for("editar_servico", servico_id=servico_id))
        db.execute(
            "UPDATE servicos SET valor_com_registro = ?, valor_sem_registro = ? WHERE id = ?",
            (v_com, v_sem, servico_id),
        )
        db.commit()
        return redirect(url_for("cadastros_servicos"))
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    return render_template("servico_editar.html", servico=servico, produtos=produtos)


@app.route("/servicos/<int:servico_id>/excluir", methods=["POST"])
def excluir_servico(servico_id):
    db = get_db()
    em_uso = db.execute(
        "SELECT COUNT(*) AS n FROM item_servicos_remessa WHERE servico_id = ?", (servico_id,)
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: este serviço já tem remessas registradas.", "erro")
    else:
        db.execute("DELETE FROM servicos WHERE id = ?", (servico_id,))
        db.commit()
    return redirect(url_for("cadastros_servicos"))


# ---------------------------------------------------------------------------
# Remessas
# ---------------------------------------------------------------------------

def itens_da_remessa(db, remessa_id):
    itens = db.execute(
        """SELECT itens_remessa.*,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.codigo AS cor_codigo, cores_estampas.descricao AS cor_descricao,
                  cores_estampas.foto AS cor_foto,
                  terceirizados.registrado AS terceirizado_registrado
           FROM itens_remessa
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           WHERE itens_remessa.remessa_id = ?
           ORDER BY (itens_remessa.prioridade IS NULL), itens_remessa.prioridade, itens_remessa.id""",
        (remessa_id,),
    ).fetchall()
    itens_view = []
    for item in itens:
        retornado = qtd_retornada(db, item["id"])
        registrado = item["terceirizado_registrado"]
        svcs = db.execute(
            """SELECT isr.id AS isr_id, isr.servico_id,
                      s.descricao AS servico_descricao,
                      s.valor_com_registro, s.valor_sem_registro
               FROM item_servicos_remessa isr
               JOIN servicos s ON s.id = isr.servico_id
               WHERE isr.item_remessa_id = ?
               ORDER BY s.descricao""",
            (item["id"],),
        ).fetchall()
        svcs_view = []
        subtotal = 0
        for svc in svcs:
            valor = svc["valor_com_registro"] if registrado else svc["valor_sem_registro"]
            sub = item["qtd_enviada"] * valor
            subtotal += sub
            svcs_view.append({**dict(svc), "servico_valor": valor, "subtotal_servico": sub})
        servico_descricao_concat = ", ".join(s["servico_descricao"] for s in svcs_view) if svcs_view else "—"
        itens_view.append({
            **dict(item),
            "produto_completo": f"{item['produto_codigo']}-{item['cor_descricao']}",
            "qtd_retornada": retornado,
            "pendente": item["qtd_enviada"] - retornado,
            "servicos": svcs_view,
            "subtotal": subtotal,
            "servico_descricao": servico_descricao_concat,
            "servico_valor": subtotal / item["qtd_enviada"] if item["qtd_enviada"] else 0,
        })
    return itens_view


@app.route("/remessas")
def remessas():
    db = get_db()
    terceirizados = db.execute("SELECT * FROM terceirizados ORDER BY nome").fetchall()
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    cores_estampas = db.execute("SELECT * FROM cores_estampas ORDER BY descricao").fetchall()
    servicos = db.execute("SELECT * FROM servicos ORDER BY descricao").fetchall()

    filtro_terceirizado_id = request.args.get("filtro_terceirizado_id", "").strip()
    filtro_data_inicio = request.args.get("filtro_data_inicio", "").strip()
    filtro_data_fim = request.args.get("filtro_data_fim", "").strip()
    buscou = bool(request.args)

    remessas_view = []
    if buscou:
        query = """SELECT remessas.*, terceirizados.nome AS terceirizado_nome
                   FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
                   WHERE 1=1"""
        params = []
        if filtro_terceirizado_id:
            query += " AND remessas.terceirizado_id = ?"
            params.append(filtro_terceirizado_id)
        if filtro_data_inicio:
            query += " AND remessas.data_envio >= ?"
            params.append(filtro_data_inicio)
        if filtro_data_fim:
            query += " AND remessas.data_envio <= ?"
            params.append(filtro_data_fim)
        query += " ORDER BY remessas.numero DESC"

        cabecalhos = db.execute(query, params).fetchall()
        hoje = date.today()
        remessas_view = []
        for r in cabecalhos:
            itens = itens_da_remessa(db, r["id"])
            tem_pendente = any(i["pendente"] > 0 for i in itens)
            try:
                data_envio = date.fromisoformat(r["data_envio"])
                dias = (hoje - data_envio).days if tem_pendente else 0
            except (ValueError, TypeError):
                dias = 0
            total_env = sum(i["qtd_enviada"] for i in itens)
            total_ret = sum(i["qtd_retornada"] for i in itens)
            pct = round(total_ret * 100 / total_env) if total_env else 0
            pago = remessa_paga(db, r["id"])
            remessas_view.append({
                **dict(r),
                "itens": itens,
                "pago": pago,
                "tem_pagamento_parcial": (not pago) and remessa_tem_retorno_pago(db, r["id"]),
                "dias_pendente": dias,
                "pct_retornado": pct,
                "total_enviado": total_env,
                "total_retornado": total_ret,
            })

    return render_template(
        "remessas.html",
        terceirizados=terceirizados,
        produtos=produtos,
        cores_estampas=cores_estampas,
        servicos=servicos,
        remessas=remessas_view,
        buscou=buscou,
        proximo_numero=proximo_numero(db, "remessas"),
        hoje=date.today().isoformat(),
        filtro_terceirizado_id=filtro_terceirizado_id,
        filtro_data_inicio=filtro_data_inicio,
        filtro_data_fim=filtro_data_fim,
    )


@app.route("/remessas/nova", methods=["POST"])
def nova_remessa():
    db = get_db()
    produto_ids       = request.form.getlist("produto_id")
    cor_ids           = request.form.getlist("cor_estampa_id")
    quantidades       = request.form.getlist("qtd_enviada")
    prioridades       = request.form.getlist("prioridade")
    previsoes_entrega = request.form.getlist("previsao_entrega")

    itens_validos = []
    for idx, (produto_id, cor_id, qtd, prioridade) in enumerate(
            zip(produto_ids, cor_ids, quantidades, prioridades)):
        if not produto_id or not cor_id or not qtd:
            continue
        try:
            qtd_int = int(qtd)
        except ValueError:
            continue
        if qtd_int <= 0:
            continue
        svc_ids = [s for s in request.form.getlist(f"svc_id_{idx}") if s]
        if not svc_ids:
            continue
        previsao = previsoes_entrega[idx].strip() if idx < len(previsoes_entrega) else ""
        itens_validos.append((produto_id, cor_id, qtd_int,
                               int(prioridade) if prioridade else None, svc_ids,
                               previsao or None))

    if not itens_validos:
        flash("Informe ao menos um item com quantidade e pelo menos um serviço.", "erro")
        return redirect(url_for("remessas"))

    terceirizado_id = request.form.get("terceirizado_id", "").strip()
    data_envio      = request.form.get("data_envio", "").strip()
    if not terceirizado_id or not data_envio:
        flash("Informe o prestador de serviço e a data de envio.", "erro")
        return redirect(url_for("remessas"))

    numero     = proximo_numero(db, "remessas")
    observacao = request.form.get("observacao", "").strip()
    try:
        cur = db.execute(
            "INSERT INTO remessas (numero, terceirizado_id, data_envio, observacao, usuario_nome) VALUES (?, ?, ?, ?, ?)",
            (numero, terceirizado_id, data_envio, observacao, session.get("usuario_nome", "")),
        )
    except Exception:
        flash("Erro ao registrar a remessa. Tente novamente.", "erro")
        return redirect(url_for("remessas"))
    remessa_id = cur.lastrowid

    for produto_id, cor_id, qtd_int, prioridade, svc_ids, previsao in itens_validos:
        cur2 = db.execute(
            """INSERT INTO itens_remessa
               (remessa_id, produto_id, cor_estampa_id, qtd_enviada, prioridade, previsao_entrega)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (remessa_id, produto_id, cor_id, qtd_int, prioridade, previsao),
        )
        item_id = cur2.lastrowid
        for svc_id in svc_ids:
            db.execute(
                "INSERT OR IGNORE INTO item_servicos_remessa (item_remessa_id, servico_id) VALUES (?, ?)",
                (item_id, svc_id),
            )

    registrar_historico(db, "remessas", remessa_id, f"Remessa Nº {numero} criada com {len(itens_validos)} item(ns)")
    db.commit()
    return redirect(url_for("confirmacao_remessa", remessa_id=remessa_id))


@app.route("/remessas/<int:remessa_id>/confirmacao")
def confirmacao_remessa(remessa_id):
    db = get_db()
    remessa = db.execute(
        """SELECT remessas.*, terceirizados.nome AS terceirizado_nome
           FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           WHERE remessas.id = ?""",
        (remessa_id,),
    ).fetchone()
    if remessa is None:
        flash("Remessa não encontrada.", "erro")
        return redirect(url_for("remessas"))
    return render_template("remessa_confirmacao.html", remessa=remessa)


@app.route("/remessas/<int:remessa_id>/editar", methods=["GET", "POST"])
def editar_remessa(remessa_id):
    db = get_db()
    remessa = db.execute(
        """SELECT remessas.*, terceirizados.nome AS terceirizado_nome, terceirizados.registrado AS terceirizado_registrado
           FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           WHERE remessas.id = ?""",
        (remessa_id,)
    ).fetchone()
    if remessa is None:
        flash("Remessa não encontrada.", "erro")
        return redirect(url_for("remessas"))
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    cores_estampas = db.execute("SELECT * FROM cores_estampas ORDER BY descricao").fetchall()
    servicos = db.execute(
        """SELECT servicos.*, produtos.codigo AS produto_codigo
           FROM servicos JOIN produtos ON produtos.id = servicos.produto_id
           ORDER BY produtos.descricao, servicos.descricao"""
    ).fetchall()
    return render_template(
        "remessa_editar.html", remessa=remessa,
        itens=itens_da_remessa(db, remessa_id),
        produtos=produtos, cores_estampas=cores_estampas, servicos=servicos
    )


@app.route("/remessas/<int:remessa_id>/adicionar-item", methods=["POST"])
def adicionar_item_remessa(remessa_id):
    db = get_db()
    remessa = db.execute("SELECT * FROM remessas WHERE id = ?", (remessa_id,)).fetchone()
    if remessa is None:
        flash("Remessa não encontrada.", "erro")
        return redirect(url_for("remessas"))
    produto_id = request.form.get("produto_id")
    cor_id     = request.form.get("cor_estampa_id")
    qtd        = request.form.get("qtd_enviada", "0")
    prioridade = request.form.get("prioridade", "").strip()
    svc_ids    = [s for s in request.form.getlist("servico_id") if s]
    if not produto_id or not cor_id or not qtd or not svc_ids:
        flash("Preencha produto, cor, quantidade e ao menos um serviço.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=remessa_id))
    try:
        qtd_int = int(qtd)
        if qtd_int <= 0:
            raise ValueError
    except ValueError:
        flash("Quantidade inválida.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=remessa_id))
    cur = db.execute(
        """INSERT INTO itens_remessa
           (remessa_id, produto_id, cor_estampa_id, qtd_enviada, prioridade)
           VALUES (?, ?, ?, ?, ?)""",
        (remessa_id, produto_id, cor_id, qtd_int,
         int(prioridade) if prioridade else None),
    )
    item_id = cur.lastrowid
    for svc_id in svc_ids:
        db.execute(
            "INSERT OR IGNORE INTO item_servicos_remessa (item_remessa_id, servico_id) VALUES (?, ?)",
            (item_id, svc_id),
        )
    db.commit()
    return redirect(url_for("editar_remessa", remessa_id=remessa_id))


@app.route("/itens-remessa/<int:item_id>/servicos/adicionar", methods=["POST"])
def adicionar_servico_item(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM itens_remessa WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        flash("Item não encontrado.", "erro")
        return redirect(url_for("remessas"))
    if item_tem_retorno_pago(db, item_id):
        flash("Item com pagamento confirmado não pode ser alterado.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=item["remessa_id"]))
    servico_id = request.form.get("servico_id")
    if not servico_id:
        flash("Selecione um serviço.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=item["remessa_id"]))
    db.execute(
        "INSERT OR IGNORE INTO item_servicos_remessa (item_remessa_id, servico_id) VALUES (?, ?)",
        (item_id, servico_id),
    )
    db.commit()
    return redirect(url_for("editar_remessa", remessa_id=item["remessa_id"]))


@app.route("/item-servico/<int:isr_id>/excluir", methods=["POST"])
def excluir_servico_item(isr_id):
    db = get_db()
    isr = db.execute(
        """SELECT isr.*, ir.remessa_id, ir.id AS item_id
           FROM item_servicos_remessa isr
           JOIN itens_remessa ir ON ir.id = isr.item_remessa_id
           WHERE isr.id = ?""",
        (isr_id,),
    ).fetchone()
    if isr is None:
        flash("Serviço não encontrado.", "erro")
        return redirect(url_for("remessas"))
    remessa_id = isr["remessa_id"]
    item_id    = isr["item_id"]
    if item_tem_retorno_pago(db, item_id):
        flash("Item com pagamento confirmado não pode ser alterado.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=remessa_id))
    count = db.execute(
        "SELECT COUNT(*) AS n FROM item_servicos_remessa WHERE item_remessa_id = ?",
        (item_id,),
    ).fetchone()["n"]
    if count <= 1:
        flash("O item precisa ter ao menos um serviço.", "erro")
        return redirect(url_for("editar_remessa", remessa_id=remessa_id))
    db.execute("DELETE FROM item_servicos_remessa WHERE id = ?", (isr_id,))
    db.commit()
    return redirect(url_for("editar_remessa", remessa_id=remessa_id))


@app.route("/remessas/<int:remessa_id>/excluir", methods=["POST"])
def excluir_remessa(remessa_id):
    if not tem_permissao("alterar_excluir_remessa"):
        flash("Você não tem permissão para excluir remessas.", "erro")
        return redirect(url_for("remessas"))
    db = get_db()
    em_uso = db.execute(
        """SELECT COUNT(*) AS n FROM itens_retorno
           JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
           WHERE itens_remessa.remessa_id = ?""",
        (remessa_id,),
    ).fetchone()["n"]
    if em_uso > 0:
        flash("Não é possível excluir: esta remessa já tem retornos vinculados.", "erro")
    elif remessa_tem_retorno_pago(db, remessa_id):
        flash("Esta remessa não pode mais ser excluída: o pagamento do período já foi confirmado.", "erro")
    else:
        remessa = db.execute("SELECT numero FROM remessas WHERE id = ?", (remessa_id,)).fetchone()
        item_ids = [r["id"] for r in db.execute(
            "SELECT id FROM itens_remessa WHERE remessa_id = ?", (remessa_id,)).fetchall()]
        for iid in item_ids:
            db.execute("DELETE FROM item_servicos_remessa WHERE item_remessa_id = ?", (iid,))
        db.execute("DELETE FROM itens_remessa WHERE remessa_id = ?", (remessa_id,))
        db.execute("DELETE FROM remessas WHERE id = ?", (remessa_id,))
        if remessa:
            registrar_historico(db, "remessas", remessa_id, f"Remessa Nº {remessa['numero']} excluída")
        db.commit()
    return redirect(url_for("remessas"))


@app.route("/remessas/exportar-excel")
def exportar_remessas_excel():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    filtro_terc = request.args.get("filtro_terceirizado_id", "").strip()
    filtro_ini  = request.args.get("filtro_data_inicio", "").strip()
    filtro_fim  = request.args.get("filtro_data_fim", "").strip()

    query = """
        SELECT remessas.*, terceirizados.nome AS terceirizado_nome, terceirizados.codigo AS terceirizado_codigo
        FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
        WHERE 1=1
    """
    params = []
    if filtro_terc:
        query += " AND remessas.terceirizado_id = ?"; params.append(filtro_terc)
    if filtro_ini:
        query += " AND remessas.data_envio >= ?"; params.append(filtro_ini)
    if filtro_fim:
        query += " AND remessas.data_envio <= ?"; params.append(filtro_fim)
    query += " ORDER BY remessas.numero DESC"
    remessas_rows = db.execute(query, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remessas"
    hf = PatternFill("solid", fgColor="F9E9DC")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.append(["REMESSAS — Casa Sanchez"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    cols = ["Nº Remessa", "Terceirizado", "Data Envio", "Status Pgto", "Observação",
            "Produto", "Cor/Estampa", "Serviço", "Qtd Enviada", "Qtd Retornada", "Pendente"]
    ws.append(cols)
    for cell in ws[3]:
        cell.font = bold; cell.fill = hf; cell.alignment = center

    for r in remessas_rows:
        itens = itens_da_remessa(db, r["id"])
        pago = remessa_paga(db, r["id"])
        for item in itens:
            ws.append([
                r["numero"],
                f"{r['terceirizado_codigo']} - {r['terceirizado_nome']}",
                r["data_envio"],
                "PAGO" if pago else "NÃO PAGO",
                r["observacao"],
                item["produto_descricao"],
                item["cor_descricao"],
                item["servico_descricao"],
                item["qtd_enviada"],
                item["qtd_retornada"],
                item["pendente"],
            ])

    col_widths = [12, 28, 14, 12, 30, 28, 18, 28, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name="remessas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/remessas/<int:remessa_id>/imprimir")
def imprimir_remessa(remessa_id):
    db = get_db()
    remessa = db.execute(
        """SELECT remessas.*, terceirizados.nome AS terceirizado_nome, terceirizados.codigo AS terceirizado_codigo
           FROM remessas JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           WHERE remessas.id = ?""",
        (remessa_id,),
    ).fetchone()
    if remessa is None:
        flash("Remessa não encontrada.", "erro")
        return redirect(url_for("remessas"))
    mps = db.execute("""
        SELECT mp.codigo, mp.descricao,
               SUM(pc.quantidade * ir.qtd_enviada) AS total
        FROM itens_remessa ir
        JOIN produto_composicao pc ON pc.produto_id = ir.produto_id
        JOIN materias_primas mp ON mp.id = pc.materia_prima_id
        WHERE ir.remessa_id = ?
        GROUP BY mp.id
        ORDER BY mp.codigo
    """, (remessa_id,)).fetchall()
    return render_template("remessa_imprimir.html", remessa=remessa, itens=itens_da_remessa(db, remessa_id), mps=mps)


# ---------------------------------------------------------------------------
# Itens de remessa (linhas individuais dentro de uma remessa)
# ---------------------------------------------------------------------------

@app.route("/itens-remessa/<int:item_id>")
def ver_item_remessa(item_id):
    db = get_db()
    item = db.execute(
        """SELECT itens_remessa.*, remessas.numero AS remessa_numero, remessas.data_envio AS data_envio,
                  terceirizados.nome AS terceirizado_nome,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.codigo AS cor_codigo, cores_estampas.descricao AS cor_descricao,
                  GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
           FROM itens_remessa
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN terceirizados ON terceirizados.id = remessas.terceirizado_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           WHERE itens_remessa.id = ?
           GROUP BY itens_remessa.id""",
        (item_id,),
    ).fetchone()
    if item is None:
        flash("Item de remessa não encontrado.", "erro")
        return redirect(url_for("remessas"))
    retornos_do_item = db.execute(
        """SELECT itens_retorno.*, retornos.numero AS retorno_numero, retornos.data_retorno
           FROM itens_retorno
           JOIN retornos ON retornos.id = itens_retorno.retorno_id
           WHERE itens_retorno.item_remessa_id = ?""",
        (item_id,),
    ).fetchall()
    retornado = qtd_retornada(db, item_id)
    return render_template(
        "item_remessa_ver.html",
        item=item,
        produto_completo=f"{item['produto_codigo']}-{item['cor_descricao']}",
        retornos=retornos_do_item,
        retornado=retornado,
        pendente=item["qtd_enviada"] - retornado,
    )


@app.route("/itens-remessa/<int:item_id>/editar", methods=["GET", "POST"])
def editar_item_remessa(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM itens_remessa WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        flash("Item de remessa não encontrado.", "erro")
        return redirect(url_for("remessas"))
    if item_tem_retorno_pago(db, item_id):
        flash("Este item não pode mais ser alterado: o pagamento do período já foi confirmado.", "erro")
        return redirect(url_for("remessas"))
    retornado = qtd_retornada(db, item_id)
    if request.method == "POST":
        qtd_enviada = int(request.form.get("qtd_enviada", 0))
        novo_produto_id = int(request.form["produto_id"])
        novo_cor_id = int(request.form["cor_estampa_id"])
        muda_item = (
            novo_produto_id != item["produto_id"]
            or novo_cor_id != item["cor_estampa_id"]
        )
        if muda_item and retornado > 0:
            flash(
                "Não é possível alterar o Produto, Cor/Estampa ou Serviço deste item: já existem retornos lançados para ele.",
                "erro",
            )
            return render_template(
                "item_remessa_editar.html",
                item=item,
                produtos=db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall(),
                cores_estampas=db.execute("SELECT * FROM cores_estampas ORDER BY descricao").fetchall(),
                servicos=db.execute("SELECT * FROM servicos ORDER BY descricao").fetchall(),
                total_itens_remessa=db.execute(
                    "SELECT COUNT(*) AS n FROM itens_remessa WHERE remessa_id = ?", (item["remessa_id"],)
                ).fetchone()["n"],
            )
        if qtd_enviada < retornado:
            flash(
                f"A quantidade enviada não pode ser menor do que o que já foi retornado neste item ({retornado}).",
                "erro",
            )
        elif qtd_enviada > 0:
            prioridade = request.form.get("prioridade", "").strip()
            previsao_entrega = request.form.get("previsao_entrega", "").strip() or None
            db.execute(
                """UPDATE itens_remessa SET produto_id = ?, cor_estampa_id = ?, qtd_enviada = ?, prioridade = ?, previsao_entrega = ?
                   WHERE id = ?""",
                (
                    request.form["produto_id"],
                    request.form["cor_estampa_id"],
                    qtd_enviada,
                    int(prioridade) if prioridade else None,
                    previsao_entrega,
                    item_id,
                ),
            )
            recalcular_finalizada(db, item_id)
            remessa_numero = db.execute(
                "SELECT numero FROM remessas WHERE id = ?", (item["remessa_id"],)
            ).fetchone()["numero"]
            registrar_historico(db, "remessas", item["remessa_id"], f"Item alterado na Remessa Nº {remessa_numero}")
            db.commit()
            return redirect(url_for("editar_remessa", remessa_id=item["remessa_id"]))
    produtos = db.execute("SELECT * FROM produtos ORDER BY descricao").fetchall()
    cores_estampas = db.execute("SELECT * FROM cores_estampas ORDER BY descricao").fetchall()
    servicos = db.execute("SELECT * FROM servicos ORDER BY descricao").fetchall()
    total_itens_remessa = db.execute(
        "SELECT COUNT(*) AS n FROM itens_remessa WHERE remessa_id = ?", (item["remessa_id"],)
    ).fetchone()["n"]
    return render_template(
        "item_remessa_editar.html",
        item=item,
        produtos=produtos,
        cores_estampas=cores_estampas,
        servicos=servicos,
        total_itens_remessa=total_itens_remessa,
    )


@app.route("/itens-remessa/<int:item_id>/excluir", methods=["POST"])
def excluir_item_remessa(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM itens_remessa WHERE id = ?", (item_id,)).fetchone()
    if item is None:
        return redirect(url_for("remessas"))
    remessa_id = item["remessa_id"]
    em_uso = qtd_retornada(db, item_id) > 0
    if em_uso:
        flash("Não é possível excluir: já existem retornos vinculados a este item.", "erro")
    elif item_tem_retorno_pago(db, item_id):
        flash("Este item não pode mais ser excluído: o pagamento do período já foi confirmado.", "erro")
    else:
        remessa_numero = db.execute(
            "SELECT numero FROM remessas WHERE id = ?", (remessa_id,)
        ).fetchone()["numero"]
        db.execute("DELETE FROM item_servicos_remessa WHERE item_remessa_id = ?", (item_id,))
        db.execute("DELETE FROM itens_remessa WHERE id = ?", (item_id,))
        restantes = db.execute(
            "SELECT COUNT(*) AS n FROM itens_remessa WHERE remessa_id = ?", (remessa_id,)
        ).fetchone()["n"]
        if restantes == 0:
            db.execute("DELETE FROM remessas WHERE id = ?", (remessa_id,))
            registrar_historico(
                db, "remessas", remessa_id,
                f"Item removido e Remessa Nº {remessa_numero} excluída (ficou sem itens)"
            )
            db.commit()
            return redirect(url_for("remessas"))
        registrar_historico(db, "remessas", remessa_id, f"Item removido da Remessa Nº {remessa_numero}")
        db.commit()
    return redirect(url_for("editar_remessa", remessa_id=remessa_id))


# ---------------------------------------------------------------------------
# Retornos
# ---------------------------------------------------------------------------

def listar_remessas_abertas(db):
    """Lista os itens de remessa (Produto+Cor/Estampa+Serviço) com saldo pendente."""
    linhas = db.execute(
        """SELECT itens_remessa.id, itens_remessa.remessa_id, itens_remessa.qtd_enviada,
                  remessas.numero AS remessa_numero, remessas.terceirizado_id AS terceirizado_id,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.descricao AS cor_descricao,
                  cores_estampas.foto AS cor_foto,
                  GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
           FROM itens_remessa
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           GROUP BY itens_remessa.id
           ORDER BY remessas.numero, itens_remessa.id"""
    ).fetchall()

    resultado = []
    for l in linhas:
        pendente = l["qtd_enviada"] - qtd_retornada(db, l["id"])
        if pendente > 0:
            resultado.append({**dict(l), "pendente": pendente})
    return resultado


@app.route("/retornos")
def retornos():
    db = get_db()
    remessas_abertas = listar_remessas_abertas(db)
    terceirizados_com_pendencia_ids = {r["terceirizado_id"] for r in remessas_abertas}
    todos_terceirizados = db.execute("SELECT * FROM terceirizados ORDER BY nome").fetchall()
    terceirizados = [t for t in todos_terceirizados if t["id"] in terceirizados_com_pendencia_ids]

    filtro_terceirizado_id = request.args.get("filtro_terceirizado_id", "").strip()
    filtro_data_inicio = request.args.get("filtro_data_inicio", "").strip()
    filtro_data_fim = request.args.get("filtro_data_fim", "").strip()
    buscou = bool(request.args)

    historico_view = []
    if buscou:
        query = """SELECT retornos.*, terceirizados.nome AS terceirizado_nome
                   FROM retornos
                   JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
                   WHERE 1=1"""
        params = []
        if filtro_terceirizado_id:
            query += " AND retornos.terceirizado_id = ?"
            params.append(filtro_terceirizado_id)
        if filtro_data_inicio:
            query += " AND retornos.data_retorno >= ?"
            params.append(filtro_data_inicio)
        if filtro_data_fim:
            query += " AND retornos.data_retorno <= ?"
            params.append(filtro_data_fim)
        query += " ORDER BY retornos.numero DESC"

        historico = db.execute(query, params).fetchall()
        for ret in historico:
            itens = db.execute(
                """SELECT itens_retorno.*, remessas.numero AS remessa_numero,
                          produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                          cores_estampas.descricao AS cor_descricao,
                          GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
                   FROM itens_retorno
                   JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
                   JOIN remessas ON remessas.id = itens_remessa.remessa_id
                   JOIN produtos ON produtos.id = itens_remessa.produto_id
                   JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
                   LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
                   LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
                   WHERE itens_retorno.retorno_id = ?
                   GROUP BY itens_retorno.id""",
                (ret["id"],),
            ).fetchall()
            pago = retorno_pago(db, ret["id"])
            historico_view.append({**dict(ret), "itens": itens, "pago": pago})

    return render_template(
        "retornos.html",
        terceirizados=terceirizados,
        todos_terceirizados=todos_terceirizados,
        remessas_abertas=remessas_abertas,
        historico=historico_view,
        buscou=buscou,
        proximo_numero=proximo_numero(db, "retornos"),
        hoje=date.today().isoformat(),
        filtro_terceirizado_id=filtro_terceirizado_id,
        filtro_data_inicio=filtro_data_inicio,
        filtro_data_fim=filtro_data_fim,
    )


@app.route("/retornos/novo", methods=["POST"])
def novo_retorno():
    db = get_db()
    terceirizado_id = request.form["terceirizado_id"]
    data_retorno = request.form["data_retorno"]
    item_ids = request.form.getlist("item_remessa_id")
    quantidades = request.form.getlist("qtd_retornada")

    itens_validos = []
    alocado_no_pedido = {}
    for item_id, qtd in zip(item_ids, quantidades):
        if not item_id or not qtd:
            continue
        qtd_int = int(qtd)
        if qtd_int <= 0:
            flash("A quantidade retornada deve ser maior que zero.", "erro")
            return redirect(url_for("retornos"))
        item = db.execute(
            """SELECT itens_remessa.*, remessas.numero AS remessa_numero
               FROM itens_remessa JOIN remessas ON remessas.id = itens_remessa.remessa_id
               WHERE itens_remessa.id = ?""",
            (item_id,),
        ).fetchone()
        if item is None:
            continue
        ja_alocado = alocado_no_pedido.get(item_id, 0)
        pendente = item["qtd_enviada"] - qtd_retornada(db, item_id) - ja_alocado
        if qtd_int > pendente:
            flash(
                f"A quantidade informada para a remessa Nº {item['remessa_numero']} ({qtd_int}) "
                f"é maior do que o saldo pendente desse item ({pendente}). Retorno não registrado.",
                "erro",
            )
            return redirect(url_for("retornos"))
        alocado_no_pedido[item_id] = ja_alocado + qtd_int
        itens_validos.append((item, qtd_int))

    if not itens_validos:
        flash("Informe ao menos um item com quantidade válida para registrar o retorno.", "erro")
        return redirect(url_for("retornos"))

    numero = proximo_numero(db, "retornos")
    observacao = request.form.get("observacao", "").strip()
    cur = db.execute(
        "INSERT INTO retornos (numero, terceirizado_id, data_retorno, observacao, usuario_nome) VALUES (?, ?, ?, ?, ?)",
        (numero, terceirizado_id, data_retorno, observacao, session.get("usuario_nome", "")),
    )
    retorno_id = cur.lastrowid

    for item, qtd_int in itens_validos:
        ja_retornado = qtd_retornada(db, item["id"])
        finalizado = 1 if (ja_retornado + qtd_int) >= item["qtd_enviada"] else 0
        db.execute(
            """INSERT INTO itens_retorno (retorno_id, item_remessa_id, qtd_retornada, finalizado)
               VALUES (?, ?, ?, ?)""",
            (retorno_id, item["id"], qtd_int, finalizado),
        )
        recalcular_finalizada(db, item["id"])

    registrar_historico(db, "retornos", retorno_id, f"Retorno Nº {numero} criado com {len(itens_validos)} item(ns)")
    db.commit()
    return redirect(url_for("confirmacao_retorno", retorno_id=retorno_id))


@app.route("/retornos/<int:retorno_id>/confirmacao")
def confirmacao_retorno(retorno_id):
    db = get_db()
    retorno = db.execute(
        """SELECT retornos.*, terceirizados.nome AS terceirizado_nome
           FROM retornos JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
           WHERE retornos.id = ?""",
        (retorno_id,),
    ).fetchone()
    if retorno is None:
        flash("Retorno não encontrado.", "erro")
        return redirect(url_for("retornos"))
    return render_template("retorno_confirmacao.html", retorno=retorno)


@app.route("/retornos/<int:retorno_id>")
def ver_retorno(retorno_id):
    db = get_db()
    retorno = db.execute(
        """SELECT retornos.*, terceirizados.nome AS terceirizado_nome
           FROM retornos JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
           WHERE retornos.id = ?""",
        (retorno_id,),
    ).fetchone()
    if retorno is None:
        flash("Retorno não encontrado.", "erro")
        return redirect(url_for("retornos"))
    itens = db.execute(
        """SELECT itens_retorno.*, remessas.numero AS remessa_numero,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.descricao AS cor_descricao,
                  GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
           FROM itens_retorno
           JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           WHERE itens_retorno.retorno_id = ?
           GROUP BY itens_retorno.id""",
        (retorno_id,),
    ).fetchall()
    return render_template("retorno_ver.html", retorno=retorno, itens=itens)


@app.route("/retornos/<int:retorno_id>/editar", methods=["GET", "POST"])
def editar_retorno(retorno_id):
    if not tem_permissao("alterar_excluir_retorno"):
        flash("Você não tem permissão para editar retornos.", "erro")
        return redirect(url_for("retornos"))
    db = get_db()
    retorno = db.execute("SELECT * FROM retornos WHERE id = ?", (retorno_id,)).fetchone()
    if retorno is None:
        flash("Retorno não encontrado.", "erro")
        return redirect(url_for("retornos"))
    if retorno_pago(db, retorno_id):
        flash("Este retorno não pode mais ser alterado: o pagamento do período já foi confirmado.", "erro")
        return redirect(url_for("retornos"))
    if request.method == "POST":
        terceirizado_id = request.form.get("terceirizado_id")
        data_retorno = request.form.get("data_retorno", "").strip()
        observacao = request.form.get("observacao", "").strip()
        if terceirizado_id and data_retorno:
            db.execute(
                "UPDATE retornos SET terceirizado_id = ?, data_retorno = ?, observacao = ? WHERE id = ?",
                (terceirizado_id, data_retorno, observacao, retorno_id),
            )
            registrar_historico(db, "retornos", retorno_id, f"Retorno Nº {retorno['numero']} editado")
            db.commit()
            return redirect(url_for("retornos"))
    terceirizados = db.execute("SELECT * FROM terceirizados ORDER BY nome").fetchall()
    itens = db.execute(
        """SELECT itens_retorno.*, remessas.numero AS remessa_numero,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.descricao AS cor_descricao,
                  GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
           FROM itens_retorno
           JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           WHERE itens_retorno.retorno_id = ?
           GROUP BY itens_retorno.id""",
        (retorno_id,),
    ).fetchall()
    return render_template(
        "retorno_editar.html", retorno=retorno, terceirizados=terceirizados, itens=itens
    )


@app.route("/retornos/<int:retorno_id>/excluir", methods=["POST"])
def excluir_retorno(retorno_id):
    if not tem_permissao("alterar_excluir_retorno"):
        flash("Você não tem permissão para excluir retornos.", "erro")
        return redirect(url_for("retornos"))
    db = get_db()
    if retorno_pago(db, retorno_id):
        flash("Este retorno não pode mais ser excluído: o pagamento do período já foi confirmado.", "erro")
        return redirect(url_for("retornos"))
    retorno = db.execute("SELECT numero FROM retornos WHERE id = ?", (retorno_id,)).fetchone()
    item_ids = [
        row["item_remessa_id"]
        for row in db.execute(
            "SELECT item_remessa_id FROM itens_retorno WHERE retorno_id = ?", (retorno_id,)
        ).fetchall()
    ]
    db.execute("DELETE FROM itens_retorno WHERE retorno_id = ?", (retorno_id,))
    db.execute("DELETE FROM retornos WHERE id = ?", (retorno_id,))
    for item_id in item_ids:
        recalcular_finalizada(db, item_id)
    if retorno:
        registrar_historico(db, "retornos", retorno_id, f"Retorno Nº {retorno['numero']} excluído")
    db.commit()
    return redirect(url_for("retornos"))


@app.route("/retornos/exportar-excel")
def exportar_retornos_excel():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    filtro_terc = request.args.get("filtro_terceirizado_id", "").strip()
    filtro_ini  = request.args.get("filtro_data_inicio", "").strip()
    filtro_fim  = request.args.get("filtro_data_fim", "").strip()

    query = """
        SELECT retornos.*, terceirizados.nome AS terceirizado_nome, terceirizados.codigo AS terceirizado_codigo
        FROM retornos JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
        WHERE 1=1
    """
    params = []
    if filtro_terc:
        query += " AND retornos.terceirizado_id = ?"; params.append(filtro_terc)
    if filtro_ini:
        query += " AND retornos.data_retorno >= ?"; params.append(filtro_ini)
    if filtro_fim:
        query += " AND retornos.data_retorno <= ?"; params.append(filtro_fim)
    query += " ORDER BY retornos.numero DESC"
    retornos_rows = db.execute(query, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retornos"
    hf = PatternFill("solid", fgColor="F9E9DC")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.append(["RETORNOS — Casa Sanchez"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    cols = ["Nº Retorno", "Terceirizado", "Data Retorno", "Observação",
            "Produto", "Cor/Estampa", "Serviço", "Nº Remessa", "Qtd Retornada", "Finalizado"]
    ws.append(cols)
    for cell in ws[3]:
        cell.font = bold; cell.fill = hf; cell.alignment = center

    for r in retornos_rows:
        itens = db.execute("""
            SELECT itens_retorno.*, remessas.numero AS remessa_numero,
                   produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                   cores_estampas.descricao AS cor_descricao,
                   GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
            FROM itens_retorno
            JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
            JOIN remessas ON remessas.id = itens_remessa.remessa_id
            JOIN produtos ON produtos.id = itens_remessa.produto_id
            JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
            LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
            LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
            WHERE itens_retorno.retorno_id = ?
            GROUP BY itens_retorno.id
        """, (r["id"],)).fetchall()
        for item in itens:
            ws.append([
                r["numero"],
                f"{r['terceirizado_codigo']} - {r['terceirizado_nome']}",
                r["data_retorno"],
                r["observacao"],
                item["produto_descricao"],
                item["cor_descricao"],
                item["servico_descricao"],
                item["remessa_numero"],
                item["qtd_retornada"],
                "SIM" if item["finalizado"] else "NÃO",
            ])

    col_widths = [12, 28, 14, 30, 28, 18, 28, 14, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name="retornos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/retornos/<int:retorno_id>/imprimir")
def imprimir_retorno(retorno_id):
    db = get_db()
    retorno = db.execute(
        """SELECT retornos.*, terceirizados.nome AS terceirizado_nome, terceirizados.codigo AS terceirizado_codigo
           FROM retornos JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
           WHERE retornos.id = ?""",
        (retorno_id,),
    ).fetchone()
    if retorno is None:
        flash("Retorno não encontrado.", "erro")
        return redirect(url_for("retornos"))
    itens = db.execute(
        """SELECT itens_retorno.*,
                  remessas.numero AS remessa_numero,
                  produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                  cores_estampas.descricao AS cor_descricao,
                  cores_estampas.foto AS cor_foto,
                  GROUP_CONCAT(servicos.descricao, ', ') AS servico_descricao
           FROM itens_retorno
           JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
           JOIN remessas ON remessas.id = itens_remessa.remessa_id
           JOIN produtos ON produtos.id = itens_remessa.produto_id
           JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
           LEFT JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
           LEFT JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
           WHERE itens_retorno.retorno_id = ?
           GROUP BY itens_retorno.id""",
        (retorno_id,),
    ).fetchall()
    return render_template("retorno_imprimir.html", retorno=retorno, itens=itens)


# ---------------------------------------------------------------------------
# Fechamento mensal
# ---------------------------------------------------------------------------

def calcular_fechamento(db, mes, terceirizado_id):
    """O fechamento cruza Produto + Cor/Estampa + Serviço + Lote de Pagamento.
    Linhas do mesmo produto pagas em lotes diferentes aparecem separadas para
    permitir visualização financeira por evento de pagamento."""
    query = """SELECT itens_retorno.retorno_id AS retorno_id,
                      terceirizados.nome AS terceirizado_nome,
                      terceirizados.registrado AS terceirizado_registrado,
                      produtos.codigo AS produto_codigo, produtos.descricao AS produto_descricao,
                      cores_estampas.descricao AS cor_descricao,
                      servicos.descricao AS servico_descricao,
                      servicos.valor_com_registro, servicos.valor_sem_registro,
                      itens_retorno.qtd_retornada AS qtd,
                      pf.id AS pagamento_id, pf.data_pagamento
               FROM itens_retorno
               JOIN retornos ON retornos.id = itens_retorno.retorno_id
               JOIN itens_remessa ON itens_remessa.id = itens_retorno.item_remessa_id
               JOIN item_servicos_remessa ON item_servicos_remessa.item_remessa_id = itens_remessa.id
               JOIN servicos ON servicos.id = item_servicos_remessa.servico_id
               JOIN terceirizados ON terceirizados.id = retornos.terceirizado_id
               JOIN produtos ON produtos.id = itens_remessa.produto_id
               JOIN cores_estampas ON cores_estampas.id = itens_remessa.cor_estampa_id
               LEFT JOIN pagamentos_fechamento_retornos pfr ON pfr.retorno_id = retornos.id
               LEFT JOIN pagamentos_fechamento pf ON pf.id = pfr.pagamento_id
               WHERE strftime('%Y-%m', retornos.data_retorno) = ?"""
    params = [mes]
    if terceirizado_id:
        query += " AND terceirizados.id = ?"
        params.append(terceirizado_id)
    query += " ORDER BY terceirizados.nome, produtos.descricao"
    linhas = db.execute(query, params).fetchall()

    resumo = {}
    for l in linhas:
        preco = l["valor_com_registro"] if l["terceirizado_registrado"] else l["valor_sem_registro"]
        pagamento_id = l["pagamento_id"]
        data_pagamento = l["data_pagamento"]
        chave = (
            l["terceirizado_nome"],
            l["produto_codigo"],
            l["produto_descricao"],
            l["cor_descricao"],
            l["servico_descricao"],
            preco,
            pagamento_id,
            data_pagamento,
        )
        resumo[chave] = resumo.get(chave, 0) + l["qtd"]

    fechamento_linhas = []
    total_geral = 0
    for (terceirizado, produto_codigo, produto, cor, servico, preco, pagamento_id, data_pagamento), qtd in resumo.items():
        total = qtd * preco
        total_geral += total
        fechamento_linhas.append({
            "terceirizado": terceirizado,
            "produto": produto,
            "produto_codigo": produto_codigo,
            "cor": cor,
            "servico": servico,
            "qtd": qtd,
            "preco": preco,
            "total": total,
            "pago": pagamento_id is not None,
            "pagamento_id": pagamento_id,
            "data_pagamento": data_pagamento,
        })
    # Pendentes (data_pagamento=None → "") primeiro, depois PAGO por data cronológica
    fechamento_linhas.sort(key=lambda x: (x["terceirizado"], x["produto"], x["data_pagamento"] or ""))
    return fechamento_linhas, total_geral


@app.route("/fechamento")
def fechamento():
    db = get_db()
    mes = request.args.get("mes", date.today().strftime("%Y-%m"))
    terceirizado_id = request.args.get("terceirizado_id", "").strip()
    buscou = bool(request.args)
    terceirizados = db.execute("""
        SELECT DISTINCT terceirizados.id, terceirizados.codigo, terceirizados.nome, terceirizados.registrado
        FROM terceirizados
        JOIN retornos ON retornos.terceirizado_id = terceirizados.id
        WHERE strftime('%Y-%m', retornos.data_retorno) = ?
        ORDER BY terceirizados.nome
    """, (mes,)).fetchall()

    fechamento_linhas, total_geral = [], 0
    total_pago = total_pendente = 0
    tem_pendentes = False
    lotes_pagamento = []
    if buscou:
        fechamento_linhas, total_geral = calcular_fechamento(db, mes, terceirizado_id)
        total_pago = sum(l["total"] for l in fechamento_linhas if l["pago"])
        total_pendente = sum(l["total"] for l in fechamento_linhas if not l["pago"])
        tem_pendentes = any(not l["pago"] for l in fechamento_linhas)
        # Agrupa totais por lote de pagamento
        lotes_dict = {}
        for l in fechamento_linhas:
            pid = l["pagamento_id"]
            if pid is not None:
                if pid not in lotes_dict:
                    lotes_dict[pid] = {"id": pid, "data_pagamento": l["data_pagamento"], "total": 0}
                lotes_dict[pid]["total"] += l["total"]
        lotes_pagamento = sorted(lotes_dict.values(), key=lambda x: x["data_pagamento"])

    return render_template(
        "fechamento.html",
        mes=mes,
        terceirizado_id=terceirizado_id,
        terceirizados=terceirizados,
        linhas=fechamento_linhas,
        total_geral=total_geral,
        total_pago=total_pago,
        total_pendente=total_pendente,
        tem_pendentes=tem_pendentes,
        lotes_pagamento=lotes_pagamento,
        buscou=buscou,
    )


@app.route("/fechamento/pagar", methods=["POST"])
def marcar_pagamento():
    if not tem_permissao("alterar_excluir_fechamento"):
        flash("Você não tem permissão para confirmar pagamentos.", "erro")
        return redirect(url_for("fechamento"))
    db = get_db()
    terceirizado_id = request.form.get("terceirizado_id", "").strip()
    mes = request.form.get("mes", "").strip()
    if not terceirizado_id or not mes:
        flash("Selecione um terceirizado e o mês para marcar o pagamento.", "erro")
        return redirect(url_for("fechamento", mes=mes))
    # Sempre cria um novo lote (sem INSERT OR IGNORE)
    cur = db.execute(
        "INSERT INTO pagamentos_fechamento (terceirizado_id, mes, data_pagamento) VALUES (?, ?, ?)",
        (terceirizado_id, mes, date.today().isoformat()),
    )
    pagamento_id = cur.lastrowid
    # Inclui apenas retornos ainda não vinculados a nenhum lote
    retorno_ids = [
        r["id"]
        for r in db.execute(
            """SELECT id FROM retornos
               WHERE terceirizado_id = ? AND strftime('%Y-%m', data_retorno) = ?
               AND id NOT IN (SELECT retorno_id FROM pagamentos_fechamento_retornos)""",
            (terceirizado_id, mes),
        ).fetchall()
    ]
    for retorno_id in retorno_ids:
        db.execute(
            "INSERT OR IGNORE INTO pagamentos_fechamento_retornos (pagamento_id, retorno_id) VALUES (?, ?)",
            (pagamento_id, retorno_id),
        )
    terceirizado_nome = db.execute(
        "SELECT nome FROM terceirizados WHERE id = ?", (terceirizado_id,)
    ).fetchone()["nome"]
    registrar_historico(
        db, "fechamento", pagamento_id,
        f"Fechamento pago: {terceirizado_nome} — {mes} ({len(retorno_ids)} retorno(s))"
    )
    db.commit()
    flash(
        f"Novo lote de pagamento confirmado: {len(retorno_ids)} retorno(s) incluídos. "
        "Retornos lançados depois desta confirmação não entram automaticamente.",
        "sucesso",
    )
    return redirect(url_for("fechamento", mes=mes, terceirizado_id=terceirizado_id))


@app.route("/fechamento/desfazer-pagamento", methods=["POST"])
def desfazer_pagamento():
    if not tem_permissao("alterar_excluir_fechamento"):
        flash("Você não tem permissão para desfazer pagamentos.", "erro")
        return redirect(url_for("fechamento"))
    db = get_db()
    pagamento_id = request.form.get("pagamento_id", "").strip()
    terceirizado_id = request.form.get("terceirizado_id", "").strip()
    mes = request.form.get("mes", "").strip()
    if pagamento_id:
        terc = db.execute("SELECT nome FROM terceirizados WHERE id = ?", (terceirizado_id,)).fetchone()
        db.execute("DELETE FROM pagamentos_fechamento_retornos WHERE pagamento_id = ?", (pagamento_id,))
        db.execute("DELETE FROM pagamentos_fechamento WHERE id = ?", (pagamento_id,))
        registrar_historico(
            db, "fechamento", pagamento_id,
            f"Fechamento desfeito: {terc['nome'] if terc else terceirizado_id} — {mes}"
        )
    db.commit()
    flash("Lote de pagamento desfeito. Os retornos deste lote foram liberados novamente.", "sucesso")
    return redirect(url_for("fechamento", mes=mes, terceirizado_id=terceirizado_id))


@app.route("/fechamento/imprimir")
def imprimir_fechamento():
    db = get_db()
    mes = request.args.get("mes", date.today().strftime("%Y-%m"))
    terceirizado_id = request.args.get("terceirizado_id", "").strip()

    fechamento_linhas, total_geral = calcular_fechamento(db, mes, terceirizado_id)
    terceirizado_nome = None
    lotes_pagamento = []
    if terceirizado_id:
        t = db.execute("SELECT nome FROM terceirizados WHERE id = ?", (terceirizado_id,)).fetchone()
        terceirizado_nome = t["nome"] if t else None
        lotes_dict = {}
        for l in fechamento_linhas:
            pid = l["pagamento_id"]
            if pid is not None:
                if pid not in lotes_dict:
                    lotes_dict[pid] = {"id": pid, "data_pagamento": l["data_pagamento"], "total": 0}
                lotes_dict[pid]["total"] += l["total"]
        lotes_pagamento = sorted(lotes_dict.values(), key=lambda x: x["data_pagamento"])

    return render_template(
        "fechamento_imprimir.html",
        mes=mes,
        terceirizado_nome=terceirizado_nome,
        linhas=fechamento_linhas,
        total_geral=total_geral,
        lotes_pagamento=lotes_pagamento,
    )


@app.route("/fechamento/exportar-excel")
def exportar_fechamento_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = get_db()
    mes = request.args.get("mes", date.today().strftime("%Y-%m"))
    terceirizado_id = request.args.get("terceirizado_id", "").strip()
    linhas, total_geral = calcular_fechamento(db, mes, terceirizado_id)
    total_pago = sum(l["total"] for l in linhas if l["pago"])
    total_pendente = sum(l["total"] for l in linhas if not l["pago"])

    mes_label = brdate(mes)
    terceirizado_nome = ""
    if terceirizado_id:
        t = db.execute("SELECT nome FROM terceirizados WHERE id = ?", (terceirizado_id,)).fetchone()
        if t:
            terceirizado_nome = t["nome"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fechamento"

    header_fill = PatternFill("solid", fgColor="F9E9DC")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.append([f"FECHAMENTO MENSAL — Casa Sanchez — {mes_label}" + (f" — {terceirizado_nome}" if terceirizado_nome else "")])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    colunas = ["Terceirizado", "Código", "Produto", "Cor/Estampa", "Serviço", "Qtde", "Preço Unit. (R$)", "Total (R$)", "Status"]
    ws.append(colunas)
    for cell in ws[3]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    for l in linhas:
        ws.append([
            l["terceirizado"], l["produto_codigo"], l["produto"],
            l["cor"], l["servico"], l["qtd"],
            float(l["preco"]), float(l["total"]),
            (f"PAGO ({l['data_pagamento']})") if l["pago"] else "PENDENTE",
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "Total já pago:", float(total_pago), ""])
    ws.append(["", "", "", "", "", "", "Total pendente:", float(total_pendente), ""])
    ws.append(["", "", "", "", "", "", "Total geral:", float(total_geral), ""])
    for row in ws.iter_rows(min_row=ws.max_row - 2, max_row=ws.max_row):
        for cell in row:
            if cell.column in (7, 8):
                cell.font = bold

    col_widths = [22, 10, 28, 18, 28, 8, 18, 16, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo = f"fechamento_{mes.replace('-', '_')}" + (f"_{terceirizado_nome.replace(' ', '_')}" if terceirizado_nome else "") + ".xlsx"
    return send_file(output, as_attachment=True, download_name=nome_arquivo,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/fechamento/resumo-por-terceirizado")
def resumo_fechamento_por_terceirizado():
    db = get_db()
    mes = request.args.get("mes", date.today().strftime("%Y-%m"))
    terceirizado_id = request.args.get("terceirizado_id", "").strip()

    fechamento_linhas, _ = calcular_fechamento(db, mes, terceirizado_id)

    grupos = {}
    ordem = []
    for l in fechamento_linhas:
        if l["terceirizado"] not in grupos:
            grupos[l["terceirizado"]] = {"terceirizado": l["terceirizado"], "linhas": [], "total": 0}
            ordem.append(l["terceirizado"])
        grupos[l["terceirizado"]]["linhas"].append(l)
        grupos[l["terceirizado"]]["total"] += l["total"]
    resumos = [grupos[nome] for nome in ordem]

    return render_template("fechamento_resumo_terceirizado.html", mes=mes, resumos=resumos)


# ---------------------------------------------------------------------------
# Importação em massa via planilha Excel
# ---------------------------------------------------------------------------

@app.route("/importar-cadastros", methods=["GET", "POST"])
def importar_cadastros():
    import io
    try:
        import openpyxl
    except ImportError:
        flash("Biblioteca openpyxl não encontrada. Instale com: pip install openpyxl", "erro")
        return redirect(url_for("cadastros"))

    if request.method == "GET":
        return render_template("importar_cadastros.html")

    arquivo = request.files.get("planilha")
    if not arquivo or not arquivo.filename.endswith(".xlsx"):
        flash("Envie um arquivo .xlsx válido.", "erro")
        return redirect(url_for("importar_cadastros"))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()), data_only=True)
    except Exception as e:
        flash(f"Erro ao abrir o arquivo: {e}", "erro")
        return redirect(url_for("importar_cadastros"))

    db = get_db()
    resultados = []

    def proc(aba_nome, col_esperadas):
        """Retorna lista de dicts com os dados da aba, ignorando linha 4 (cabeçalho) e linhas em branco."""
        if aba_nome not in wb.sheetnames:
            return []
        ws = wb[aba_nome]
        rows = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row[:col_esperadas]]
            if all(v == "" for v in vals):
                continue
            if len(vals[0]) > 50:  # ignora linhas de rodapé/nota
                continue
            rows.append(vals)
        return rows

    # ── Catálogo de Cores / Estampas ─────────────────────────────────────────
    ok_cc = skip_cc = 0
    for ln, (descricao,) in enumerate(proc("0-Catalogo_Cores", 1), start=5):
        descricao = descricao.upper()
        if not descricao:
            continue
        try:
            db.execute("INSERT INTO catalogo_cores (descricao) VALUES (?)", (descricao,))
            ok_cc += 1
        except sqlite3.IntegrityError:
            skip_cc += 1
    if ok_cc or skip_cc:
        resultados.append(("info", f"Catálogo de Cores: {ok_cc} inserida(s), {skip_cc} já existia(m)."))

    # ── Catálogo de Serviços ─────────────────────────────────────────────────
    ok_cs = skip_cs = 0
    for ln, (descricao,) in enumerate(proc("0b-Catalogo_Servicos", 1), start=5):
        descricao = descricao.upper()
        if not descricao:
            continue
        try:
            db.execute("INSERT INTO catalogo_servicos (descricao) VALUES (?)", (descricao,))
            ok_cs += 1
        except sqlite3.IntegrityError:
            skip_cs += 1
    if ok_cs or skip_cs:
        resultados.append(("info", f"Catálogo de Serviços: {ok_cs} inserida(s), {skip_cs} já existia(m)."))

    # ── Produtos ─────────────────────────────────────────────────────────────
    ok_p = err_p = skip_p = 0
    for ln, (codigo, descricao) in enumerate(proc("2-Produtos", 2), start=5):
        if not codigo or not descricao:
            err_p += 1
            resultados.append(("erro", f"Produtos linha {ln}: Código e Descrição são obrigatórios."))
            continue
        try:
            db.execute("INSERT INTO produtos (codigo, descricao) VALUES (?, ?)", (codigo, descricao))
            ok_p += 1
        except sqlite3.IntegrityError:
            skip_p += 1
    if ok_p or skip_p or err_p:
        resultados.append(("info", f"Produtos: {ok_p} inserido(s), {skip_p} já existia(m), {err_p} erro(s)."))

    # ── Prestadores ──────────────────────────────────────────────────────────
    ok_t = err_t = skip_t = 0
    for ln, vals in enumerate(proc("1-Prestadores", 4), start=5):
        codigo, nome = vals[0], vals[1]
        telefone = vals[2] if len(vals) > 2 else ""
        registrado_raw = vals[3] if len(vals) > 3 else "1"
        if not codigo or not nome:
            err_t += 1
            resultados.append(("erro", f"Prestadores linha {ln}: Código e Nome são obrigatórios."))
            continue
        registrado = 0 if str(registrado_raw).strip() == "0" else 1
        try:
            db.execute("INSERT INTO terceirizados (codigo, nome, telefone, registrado) VALUES (?, ?, ?, ?)",
                       (codigo, nome, telefone or None, registrado))
            ok_t += 1
        except sqlite3.IntegrityError:
            db.execute("UPDATE terceirizados SET nome=?, telefone=?, registrado=? WHERE codigo=?",
                       (nome, telefone or None, registrado, codigo))
            skip_t += 1
    if ok_t or skip_t or err_t:
        resultados.append(("info", f"Prestadores: {ok_t} inserido(s), {skip_t} já existia(m), {err_t} erro(s)."))

    # ── Cores / Estampas ─────────────────────────────────────────────────────
    ok_c = err_c = skip_c = 0
    for ln, (codigo, produto_codigo, descricao) in enumerate(proc("3-Cores_Estampas", 3), start=5):
        if not produto_codigo and not descricao:
            continue  # linha vazia ou só com sugestão de código cinza — ignorar
        if not codigo or not produto_codigo or not descricao:
            err_c += 1
            resultados.append(("erro", f"Cores linha {ln}: todos os campos são obrigatórios."))
            continue
        prod = db.execute("SELECT id FROM produtos WHERE codigo = ?", (produto_codigo,)).fetchone()
        if not prod:
            err_c += 1
            resultados.append(("erro", f"Cores linha {ln}: produto '{produto_codigo}' não encontrado."))
            continue
        try:
            db.execute("INSERT INTO cores_estampas (codigo, produto_id, descricao) VALUES (?, ?, ?)",
                       (codigo, prod["id"], descricao))
            ok_c += 1
        except sqlite3.IntegrityError:
            skip_c += 1
    if ok_c or skip_c or err_c:
        resultados.append(("info", f"Cores/Estampas: {ok_c} inserida(s), {skip_c} já existia(m), {err_c} erro(s)."))

    # ── Serviços ─────────────────────────────────────────────────────────────
    ok_s = err_s = skip_s = 0
    for ln, vals in enumerate(proc("4-Servicos", 5), start=5):
        codigo, produto_codigo, descricao = vals[0], vals[1], vals[2]
        valor_com_raw = vals[3] if len(vals) > 3 else ""
        valor_sem_raw = vals[4] if len(vals) > 4 else ""
        if not produto_codigo and not descricao:
            continue  # linha vazia ou só com sugestão de código cinza — ignorar
        if not codigo or not produto_codigo or not descricao:
            err_s += 1
            resultados.append(("erro", f"Serviços linha {ln}: todos os campos são obrigatórios."))
            continue
        try:
            valor_com = float(str(valor_com_raw).replace(",", "."))
        except (ValueError, TypeError):
            err_s += 1
            resultados.append(("erro", f"Serviços linha {ln}: valor inválido (use número com ponto, ex: 2.50)."))
            continue
        # valor_sem_registro é opcional — se vazio usa o mesmo valor
        try:
            valor_sem = float(str(valor_sem_raw).replace(",", ".")) if valor_sem_raw else valor_com
        except (ValueError, TypeError):
            valor_sem = valor_com
        prod = db.execute("SELECT id FROM produtos WHERE codigo = ?", (produto_codigo,)).fetchone()
        if not prod:
            err_s += 1
            resultados.append(("erro", f"Serviços linha {ln}: produto '{produto_codigo}' não encontrado."))
            continue
        try:
            db.execute(
                "INSERT INTO servicos (codigo, produto_id, descricao, valor_com_registro, valor_sem_registro) VALUES (?, ?, ?, ?, ?)",
                (codigo, prod["id"], descricao, valor_com, valor_sem))
            ok_s += 1
        except sqlite3.IntegrityError:
            skip_s += 1
    if ok_s or skip_s or err_s:
        resultados.append(("info", f"Serviços: {ok_s} inserido(s), {skip_s} já existia(m), {err_s} erro(s)."))

    # ── Aba 5-Composicao_MP ────────────────────────────────────────────────
    ws_comp = wb["5-Composicao_MP"] if "5-Composicao_MP" in wb.sheetnames else None
    if ws_comp:
        ok_c = skip_c = err_c = 0
        for ln, row in enumerate(ws_comp.iter_rows(min_row=5, values_only=True), start=5):
            produto_codigo = str(row[0]).strip() if row[0] else ""
            mp_codigo      = str(row[1]).strip() if row[1] else ""
            mp_descricao   = str(row[2]).strip() if row[2] else ""
            quantidade_raw = row[3]
            if not produto_codigo or not mp_codigo or not mp_descricao or quantidade_raw is None:
                continue
            try:
                quantidade = float(str(quantidade_raw).replace(",", "."))
            except ValueError:
                err_c += 1
                resultados.append(("erro", f"Composição linha {ln}: quantidade inválida '{quantidade_raw}'."))
                continue
            prod = db.execute("SELECT id FROM produtos WHERE codigo = ?", (produto_codigo,)).fetchone()
            if not prod:
                err_c += 1
                resultados.append(("erro", f"Composição linha {ln}: produto '{produto_codigo}' não encontrado."))
                continue
            db.execute("INSERT OR IGNORE INTO materias_primas (codigo, descricao) VALUES (?, ?)", (mp_codigo, mp_descricao))
            mp = db.execute("SELECT id FROM materias_primas WHERE codigo = ?", (mp_codigo,)).fetchone()
            try:
                db.execute(
                    "INSERT INTO produto_composicao (produto_id, materia_prima_id, quantidade) VALUES (?, ?, ?)",
                    (prod["id"], mp["id"], quantidade))
                ok_c += 1
            except sqlite3.IntegrityError:
                db.execute(
                    "UPDATE produto_composicao SET quantidade = ? WHERE produto_id = ? AND materia_prima_id = ?",
                    (quantidade, prod["id"], mp["id"]))
                skip_c += 1
        if ok_c or skip_c or err_c:
            resultados.append(("info", f"Composição MP: {ok_c} inserido(s), {skip_c} atualizado(s), {err_c} erro(s)."))

    db.commit()
    return render_template("importar_cadastros.html", resultados=resultados)


# ---------------------------------------------------------------------------
# Cadastro de Matérias-Primas
# ---------------------------------------------------------------------------

@app.route("/cadastros/materias-primas")
def cadastros_materias_primas():
    db = get_db()
    mps = db.execute("SELECT * FROM materias_primas ORDER BY codigo").fetchall()
    return render_template("cadastro_materias_primas.html", mps=mps)


@app.route("/cadastros/materia-prima/adicionar", methods=["POST"])
def materia_prima_adicionar():
    db = get_db()
    codigo    = request.form.get("codigo", "").strip()
    descricao = request.form.get("descricao", "").strip()
    unidade   = request.form.get("unidade", "UN").strip().upper() or "UN"
    if not codigo or not descricao:
        flash("Informe o código e a descrição da matéria-prima.", "erro")
    else:
        try:
            db.execute(
                "INSERT INTO materias_primas (codigo, descricao, unidade) VALUES (?, ?, ?)",
                (codigo, descricao, unidade),
            )
            db.commit()
            flash("Matéria-prima adicionada.", "sucesso")
        except sqlite3.IntegrityError:
            flash(f'Já existe uma matéria-prima com o código "{codigo}".', "erro")
    return redirect(url_for("cadastros_materias_primas"))


@app.route("/cadastros/materia-prima/<int:mp_id>/editar", methods=["POST"])
def materia_prima_editar(mp_id):
    db = get_db()
    descricao = request.form.get("descricao", "").strip()
    unidade   = request.form.get("unidade", "UN").strip().upper() or "UN"
    if not descricao:
        flash("Descrição não pode ser vazia.", "erro")
        return redirect(url_for("cadastros_materias_primas"))
    db.execute("UPDATE materias_primas SET descricao = ?, unidade = ? WHERE id = ?",
               (descricao, unidade, mp_id))
    db.commit()
    flash("Matéria-prima atualizada. Todas as composições vinculadas refletem a alteração automaticamente.", "ok")
    return redirect(url_for("cadastros_materias_primas"))


# ---------------------------------------------------------------------------
# Cadastro de Fornecedores
# ---------------------------------------------------------------------------

@app.route("/cadastros/fornecedores")
def cadastros_fornecedores():
    db = get_db()
    fornecedores = db.execute("SELECT * FROM fornecedores ORDER BY codigo").fetchall()
    return render_template("cadastro_fornecedores.html", fornecedores=fornecedores)


@app.route("/cadastros/fornecedor/adicionar", methods=["POST"])
def fornecedor_adicionar():
    db = get_db()
    codigo = request.form.get("codigo", "").strip()
    nome = request.form.get("nome_razao_social", "").strip()
    if not codigo or not nome:
        flash("Código e Nome/Razão Social são obrigatórios.", "erro")
        return redirect(url_for("cadastros_fornecedores"))
    try:
        db.execute(
            "INSERT INTO fornecedores (codigo, nome_razao_social, telefone_empresa, nome_representante, telefone_representante) VALUES (?,?,?,?,?)",
            (codigo, nome,
             request.form.get("telefone_empresa", "").strip() or None,
             request.form.get("nome_representante", "").strip() or None,
             request.form.get("telefone_representante", "").strip() or None)
        )
        db.commit()
        flash("Fornecedor cadastrado.", "ok")
    except Exception:
        flash("Código já existe.", "erro")
    return redirect(url_for("cadastros_fornecedores"))


@app.route("/cadastros/fornecedor/<int:forn_id>/editar", methods=["POST"])
def fornecedor_editar(forn_id):
    db = get_db()
    nome = request.form.get("nome_razao_social", "").strip()
    if not nome:
        flash("Nome/Razão Social é obrigatório.", "erro")
        return redirect(url_for("cadastros_fornecedores"))
    db.execute(
        "UPDATE fornecedores SET nome_razao_social=?, telefone_empresa=?, nome_representante=?, telefone_representante=? WHERE id=?",
        (nome,
         request.form.get("telefone_empresa", "").strip() or None,
         request.form.get("nome_representante", "").strip() or None,
         request.form.get("telefone_representante", "").strip() or None,
         forn_id)
    )
    db.commit()
    flash("Fornecedor atualizado.", "ok")
    return redirect(url_for("cadastros_fornecedores"))


@app.route("/cadastros/fornecedor/<int:forn_id>/excluir", methods=["POST"])
def fornecedor_excluir(forn_id):
    db = get_db()
    ids_cols = ["fornecedor1_id", "fornecedor2_id", "fornecedor3_id", "fornecedor4_id", "fornecedor5_id"]
    where = " OR ".join(f"{c}=?" for c in ids_cols)
    vinculado = db.execute(f"SELECT id FROM planos_corte WHERE {where}", [forn_id] * 5).fetchone()
    if vinculado:
        flash("Este fornecedor está vinculado a um ou mais planos de corte e não pode ser excluído.", "erro")
        return redirect(url_for("cadastros_fornecedores"))
    db.execute("DELETE FROM fornecedores WHERE id=?", (forn_id,))
    db.commit()
    flash("Fornecedor excluído.", "ok")
    return redirect(url_for("cadastros_fornecedores"))


# ---------------------------------------------------------------------------
# Cronometragem de Corte
# ---------------------------------------------------------------------------

def _format_hhmm(total_min):
    if total_min is None:
        return ""
    h = int(total_min) // 60
    m = int(total_min) % 60
    return f"{h:02d}:{m:02d}"


@app.route("/produtos/<int:produto_id>/cronometragem", methods=["GET", "POST"])
def produto_cronometragem(produto_id):
    db = get_db()
    produto = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not produto:
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("cadastros_produtos"))

    if request.method == "POST":
        for n in (1, 2, 3):
            hora_inicio = request.form.get(f"hora_inicio_{n}", "").strip() or None
            hora_fim    = request.form.get(f"hora_fim_{n}",    "").strip() or None
            qtd         = request.form.get(f"qtd_{n}",         "").strip() or None
            pessoas     = request.form.get(f"pessoas_{n}",     "").strip() or None
            data_med    = request.form.get(f"data_{n}",        "").strip() or None
            nomes       = request.form.get(f"nomes_{n}",       "").strip() or None
            t_h, t_m = 0, 0
            if hora_inicio and hora_fim:
                try:
                    hi_h, hi_m = map(int, hora_inicio.split(":"))
                    hf_h, hf_m = map(int, hora_fim.split(":"))
                    total_min = (hf_h * 60 + hf_m) - (hi_h * 60 + hi_m)
                    if total_min < 0:
                        total_min += 24 * 60  # atravessou meia-noite
                    t_h, t_m = divmod(total_min, 60)
                except ValueError:
                    pass
            db.execute("""
                INSERT INTO cronometragem_corte
                    (produto_id, numero, tempo_horas, tempo_minutos, qtd_cortada, num_pessoas,
                     data_medicao, nomes_pessoas, hora_inicio, hora_fim)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(produto_id, numero) DO UPDATE SET
                    tempo_horas=excluded.tempo_horas, tempo_minutos=excluded.tempo_minutos,
                    qtd_cortada=excluded.qtd_cortada, num_pessoas=excluded.num_pessoas,
                    data_medicao=excluded.data_medicao, nomes_pessoas=excluded.nomes_pessoas,
                    hora_inicio=excluded.hora_inicio, hora_fim=excluded.hora_fim
            """, (produto_id, n, t_h, t_m, qtd or None, pessoas or None, data_med, nomes,
                  hora_inicio, hora_fim))
        db.commit()
        flash("Cronometragem salva.", "ok")
        return redirect(url_for("produto_cronometragem", produto_id=produto_id))

    rows = {r["numero"]: r for r in db.execute(
        "SELECT * FROM cronometragem_corte WHERE produto_id = ? ORDER BY numero", (produto_id,)
    ).fetchall()}

    medidas = []
    for n in (1, 2, 3):
        r = rows.get(n)
        total_min = (r["tempo_horas"] * 60 + r["tempo_minutos"]) if r else 0
        medidas.append({
            "numero":       n,
            "tempo_str":    _format_hhmm(total_min) if (r and (r["tempo_horas"] or r["tempo_minutos"])) else "",
            "qtd_cortada":  r["qtd_cortada"]  if r else "",
            "num_pessoas":  r["num_pessoas"]  if r else "",
            "data_medicao": r["data_medicao"] if r else "",
            "nomes_pessoas": r["nomes_pessoas"] or "" if r else "",
            "hora_inicio":  r["hora_inicio"]  or "" if r else "",
            "hora_fim":     r["hora_fim"]     or "" if r else "",
        })

    media = _calcular_media_cronometragem(medidas)
    return render_template("produto_cronometragem.html", produto=produto,
                           medidas=medidas, media=media)


@app.route("/produtos/cronometragem/importar-excel", methods=["POST"])
def importar_cronometragem_excel():
    import io, openpyxl
    from datetime import datetime, date as dt_date, time as dt_time

    f = request.files.get("arquivo")
    if not f or not f.filename.endswith(".xlsx"):
        flash("Selecione um arquivo .xlsx válido.", "erro")
        return redirect(request.referrer or url_for("cadastros_produtos"))

    db = get_db()
    produtos_map = {r["codigo"]: r["id"] for r in db.execute("SELECT id, codigo FROM produtos").fetchall()}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception:
        flash("Arquivo inválido ou corrompido.", "erro")
        return redirect(request.referrer or url_for("cadastros_produtos"))

    ws = wb.active
    atualizados, erros = 0, []

    def _para_hora_str(val):
        if isinstance(val, dt_time):
            return val.strftime("%H:%M")
        if isinstance(val, float):
            total = round(val * 24 * 60)
            return f"{total // 60:02d}:{total % 60:02d}"
        if isinstance(val, str) and val.strip():
            return val.strip()[:5]
        return None

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        codigo   = str(row[0]).strip() if row[0] is not None else ""
        num_raw  = row[2]
        if not codigo or num_raw is None:
            continue
        try:
            numero = int(num_raw)
        except (ValueError, TypeError):
            continue
        if numero not in (1, 2, 3):
            continue
        produto_id = produtos_map.get(codigo)
        if not produto_id:
            erros.append(f"Linha {i}: código '{codigo}' não encontrado")
            continue

        # Data
        data_str = None
        dv = row[3]
        if isinstance(dv, (datetime, dt_date)):
            d = dv.date() if isinstance(dv, datetime) else dv
            data_str = d.strftime("%Y-%m-%d")
        elif isinstance(dv, str) and dv.strip():
            try:
                d2, m2, a2 = dv.strip().split("/")
                data_str = f"{a2}-{m2}-{d2}"
            except Exception:
                pass

        hora_inicio = _para_hora_str(row[4])
        hora_fim    = _para_hora_str(row[5])

        # Calcular tempo a partir de hora_inicio e hora_fim
        t_h, t_m = 0, 0
        if hora_inicio and hora_fim:
            try:
                hi_h, hi_m = map(int, hora_inicio.split(":"))
                hf_h, hf_m = map(int, hora_fim.split(":"))
                diff = (hf_h * 60 + hf_m) - (hi_h * 60 + hi_m)
                if diff < 0:
                    diff += 24 * 60
                t_h, t_m = divmod(diff, 60)
            except Exception:
                pass
        else:
            # Fallback: usar coluna Tempo se preenchida
            tv = row[6]
            if isinstance(tv, dt_time):
                t_h, t_m = tv.hour, tv.minute
            elif isinstance(tv, float):
                total = round(tv * 24 * 60)
                t_h, t_m = divmod(total, 60)
            elif isinstance(tv, str) and tv.strip():
                try:
                    t_h, t_m = map(int, tv.strip().split(":"))
                except Exception:
                    pass

        qtd    = int(row[7]) if row[7] is not None else None
        pessoas = int(row[8]) if row[8] is not None else None
        nomes  = str(row[9]).strip() if row[9] is not None else None

        db.execute("""
            INSERT INTO cronometragem_corte
                (produto_id, numero, tempo_horas, tempo_minutos, qtd_cortada, num_pessoas,
                 data_medicao, nomes_pessoas, hora_inicio, hora_fim)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(produto_id, numero) DO UPDATE SET
                tempo_horas=excluded.tempo_horas, tempo_minutos=excluded.tempo_minutos,
                qtd_cortada=excluded.qtd_cortada, num_pessoas=excluded.num_pessoas,
                data_medicao=excluded.data_medicao, nomes_pessoas=excluded.nomes_pessoas,
                hora_inicio=excluded.hora_inicio, hora_fim=excluded.hora_fim
        """, (produto_id, numero, t_h, t_m, qtd, pessoas, data_str, nomes, hora_inicio, hora_fim))
        atualizados += 1

    db.commit()
    msg = f"{atualizados} medição(ões) importada(s) com sucesso."
    if erros:
        flash(msg + " Atenção: " + "; ".join(erros[:5]), "erro")
    else:
        flash(msg, "ok")
    return redirect(request.referrer or url_for("cadastros_produtos"))


@app.route("/produtos/cronometragem/exportar-excel")
def exportar_cronometragem_excel():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Protection
    from openpyxl.styles.fills import PatternFill
    from datetime import date as dt_date, time as dt_time

    def _parse_date(s):
        if not s:
            return None
        try:
            a, m, d = s.split("-")
            return dt_date(int(a), int(m), int(d))
        except Exception:
            return None

    def _parse_time(s):
        if not s:
            return None
        try:
            h, m = s.split(":")
            return dt_time(int(h), int(m))
        except Exception:
            return None

    db = get_db()
    produtos = db.execute("SELECT id, codigo, descricao FROM produtos ORDER BY codigo").fetchall()
    rows_por_produto = {}
    for row in db.execute("SELECT * FROM cronometragem_corte ORDER BY produto_id, numero").fetchall():
        rows_por_produto.setdefault(row["produto_id"], {})[row["numero"]] = row

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronometragem"

    fill_header  = PatternFill("solid", fgColor="F9E9DC")
    fill_tempo   = PatternFill("solid", fgColor="EDF4E8")  # verde claro = calculado
    bold         = Font(bold=True)
    bold_verde   = Font(bold=True, color="2E6B1F")
    center       = Alignment(horizontal="center", vertical="center")

    cabecalho = ["Código", "Produto", "Medição Nº", "Data", "Hora Inicial", "Hora Final", "Tempo (HH:MM)", "Qtde Cortada", "Nº Pessoas", "Nomes dos Cortadores"]
    for col, titulo in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = bold
        c.fill = fill_header
        c.alignment = center
    # Cabeçalho Tempo com destaque verde
    ws.cell(row=1, column=7).font = bold_verde

    linha = 2
    for p in produtos:
        cron = rows_por_produto.get(p["id"], {})
        for n in (1, 2, 3):
            r = cron.get(n)

            # Coluna A - Código
            ws.cell(row=linha, column=1, value=p["codigo"])
            # Coluna B - Produto
            ws.cell(row=linha, column=2, value=p["descricao"])
            # Coluna C - Medição Nº
            c = ws.cell(row=linha, column=3, value=n)
            c.alignment = center
            # Coluna D - Data (tipo data, formato DD/MM/YYYY)
            c = ws.cell(row=linha, column=4, value=_parse_date(r["data_medicao"] if r else None))
            c.number_format = "DD/MM/YYYY"
            c.alignment = center
            # Coluna E - Hora Inicial (tipo hora, formato HH:MM)
            c = ws.cell(row=linha, column=5, value=_parse_time(r["hora_inicio"] if r else None))
            c.number_format = "HH:MM"
            c.alignment = center
            # Coluna F - Hora Final (tipo hora, formato HH:MM)
            c = ws.cell(row=linha, column=6, value=_parse_time(r["hora_fim"] if r else None))
            c.number_format = "HH:MM"
            c.alignment = center
            # Coluna G - Tempo calculado: fórmula MOD(F-E,1) trata meia-noite
            c = ws.cell(row=linha, column=7,
                        value=f'=IF(AND(E{linha}<>"",F{linha}<>""),MOD(F{linha}-E{linha},1),"")')
            c.number_format = "HH:MM"
            c.alignment = center
            c.fill = fill_tempo
            c.font = Font(bold=True, color="2E6B1F")
            # Coluna H - Qtde Cortada
            ws.cell(row=linha, column=8,
                    value=r["qtd_cortada"] if r and r["qtd_cortada"] else None)
            # Coluna I - Nº Pessoas
            ws.cell(row=linha, column=9,
                    value=r["num_pessoas"] if r and r["num_pessoas"] else None)
            # Coluna J - Nomes
            ws.cell(row=linha, column=10,
                    value=r["nomes_pessoas"] if r and r["nomes_pessoas"] else None)

            linha += 1

    col_widths = [14, 36, 12, 14, 13, 13, 15, 14, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="cronometragem_corte.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _calcular_media_cronometragem(medidas):
    validas = [m for m in medidas if m["tempo_str"]]
    if not validas:
        return None
    total_min = 0
    total_qtd = 0
    total_pessoas = 0
    n_qtd = 0
    n_pessoas = 0
    for m in validas:
        parts = m["tempo_str"].split(":")
        total_min += int(parts[0]) * 60 + int(parts[1])  # HH:MM → total minutes
        if m["qtd_cortada"]:
            total_qtd += int(m["qtd_cortada"])
            n_qtd += 1
        if m["num_pessoas"]:
            total_pessoas += int(m["num_pessoas"])
            n_pessoas += 1
    count = len(validas)
    avg_min = total_min / count
    avg_qtd = round(total_qtd / n_qtd, 1) if n_qtd else None
    avg_pessoas = round(total_pessoas / n_pessoas, 1) if n_pessoas else None
    pcs_hora = round((avg_qtd / (avg_min / 60)), 1) if (avg_qtd and avg_min) else None
    return {
        "tempo_str": _format_hhmm(round(avg_min)),
        "qtd_cortada": avg_qtd,
        "num_pessoas": avg_pessoas,
        "pcs_hora": pcs_hora,
    }


# ---------------------------------------------------------------------------
# Histórico de alterações
# ---------------------------------------------------------------------------

@app.route("/historico")
def historico():
    db = get_db()
    filtro_tabela = request.args.get("filtro_tabela", "").strip()
    filtro_usuario = request.args.get("filtro_usuario", "").strip()
    filtro_data_inicio = request.args.get("filtro_data_inicio", "").strip()
    filtro_data_fim = request.args.get("filtro_data_fim", "").strip()

    query = "SELECT * FROM historico WHERE 1=1"
    params = []
    if filtro_tabela:
        query += " AND tabela = ?"
        params.append(filtro_tabela)
    if filtro_usuario:
        query += " AND usuario_nome LIKE ?"
        params.append(f"%{filtro_usuario}%")
    if filtro_data_inicio:
        query += " AND date(data_hora) >= ?"
        params.append(filtro_data_inicio)
    if filtro_data_fim:
        query += " AND date(data_hora) <= ?"
        params.append(filtro_data_fim)
    query += " ORDER BY id DESC LIMIT 500"

    registros = db.execute(query, params).fetchall()
    tabelas_disponiveis = [r["tabela"] for r in db.execute(
        "SELECT DISTINCT tabela FROM historico ORDER BY tabela"
    ).fetchall()]
    return render_template(
        "historico.html",
        registros=registros,
        tabelas_disponiveis=tabelas_disponiveis,
        filtro_tabela=filtro_tabela,
        filtro_usuario=filtro_usuario,
        filtro_data_inicio=filtro_data_inicio,
        filtro_data_fim=filtro_data_fim,
        hoje=date.today().isoformat(),
    )


# ---------------------------------------------------------------------------
# Composição de Produtos (Ficha Técnica de MP)
# ---------------------------------------------------------------------------

@app.route("/produtos/<int:produto_id>/composicao", methods=["GET", "POST"])
def produto_composicao(produto_id):
    db = get_db()
    produto = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not produto:
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("cadastros_produtos"))

    if request.method == "POST":
        acao = request.form.get("acao")

        if acao == "adicionar":
            mp_id_str  = request.form.get("mp_id", "").strip()
            qtd_str    = request.form.get("quantidade", "").strip().replace(",", ".")
            try:
                mp_id_add = int(mp_id_str)
                qtd_add   = float(qtd_str)
                if qtd_add <= 0:
                    raise ValueError
            except ValueError:
                flash("Quantidade inválida.", "erro")
                return redirect(url_for("produto_composicao", produto_id=produto_id))
            try:
                db.execute(
                    "INSERT INTO produto_composicao (produto_id, materia_prima_id, quantidade) VALUES (?,?,?)",
                    (produto_id, mp_id_add, qtd_add))
                db.commit()
                flash("Matéria-prima adicionada à composição.", "ok")
            except sqlite3.IntegrityError:
                flash("Essa matéria-prima já está na composição.", "erro")

        elif acao == "editar":
            mp_id_str = request.form.get("mp_id", "").strip()
            qtd_str   = request.form.get("quantidade", "").strip().replace(",", ".")
            try:
                mp_id_ed = int(mp_id_str)
                qtd_ed   = float(qtd_str)
                if qtd_ed <= 0:
                    raise ValueError
            except ValueError:
                flash("Quantidade inválida.", "erro")
                return redirect(url_for("produto_composicao", produto_id=produto_id))
            db.execute(
                "UPDATE produto_composicao SET quantidade = ? WHERE produto_id = ? AND materia_prima_id = ?",
                (qtd_ed, produto_id, mp_id_ed))
            db.commit()
            flash("Quantidade atualizada.", "ok")

        elif acao == "remover":
            mp_id_str = request.form.get("mp_id", "").strip()
            try:
                mp_id_rm = int(mp_id_str)
            except ValueError:
                flash("Item inválido.", "erro")
                return redirect(url_for("produto_composicao", produto_id=produto_id))
            db.execute(
                "DELETE FROM produto_composicao WHERE produto_id = ? AND materia_prima_id = ?",
                (produto_id, mp_id_rm))
            db.commit()
            flash("Item removido da composição.", "ok")

        return redirect(url_for("produto_composicao", produto_id=produto_id))

    itens = db.execute("""
        SELECT mp.id AS mp_id, mp.codigo, mp.descricao, mp.unidade, pc.quantidade
        FROM produto_composicao pc
        JOIN materias_primas mp ON pc.materia_prima_id = mp.id
        WHERE pc.produto_id = ?
        ORDER BY mp.codigo
    """, (produto_id,)).fetchall()
    todas_mps = db.execute(
        "SELECT id, codigo, descricao, unidade FROM materias_primas ORDER BY codigo"
    ).fetchall()
    return render_template("produto_composicao.html", produto=produto,
                           itens=itens, todas_mps=todas_mps)


@app.route("/remessas/<int:remessa_id>/materias-primas")
def remessa_materias_primas(remessa_id):
    db = get_db()
    remessa = db.execute("""
        SELECT r.*, t.nome AS terceirizado_nome, t.codigo AS terceirizado_codigo
        FROM remessas r JOIN terceirizados t ON r.terceirizado_id = t.id
        WHERE r.id = ?
    """, (remessa_id,)).fetchone()
    if not remessa:
        flash("Remessa não encontrada.", "erro")
        return redirect(url_for("remessas"))
    # Calcula totais de MP: soma quantidade_ficha * qtd_enviada para cada MP
    mps = db.execute("""
        SELECT mp.codigo, mp.descricao, mp.unidade,
               SUM(pc.quantidade * ir.qtd_enviada) AS total
        FROM itens_remessa ir
        JOIN produto_composicao pc ON pc.produto_id = ir.produto_id
        JOIN materias_primas mp ON mp.id = pc.materia_prima_id
        WHERE ir.remessa_id = ?
        GROUP BY mp.id
        ORDER BY mp.codigo
    """, (remessa_id,)).fetchall()
    # Itens da remessa para contexto
    itens = db.execute("""
        SELECT p.codigo AS produto_codigo, p.descricao AS produto_descricao,
               ce.descricao AS cor_descricao,
               GROUP_CONCAT(s.descricao, ', ') AS servico_descricao,
               ir.qtd_enviada
        FROM itens_remessa ir
        JOIN produtos p ON ir.produto_id = p.id
        JOIN cores_estampas ce ON ir.cor_estampa_id = ce.id
        LEFT JOIN item_servicos_remessa isr ON isr.item_remessa_id = ir.id
        LEFT JOIN servicos s ON s.id = isr.servico_id
        WHERE ir.remessa_id = ?
        GROUP BY ir.id
        ORDER BY ir.prioridade, ir.id
    """, (remessa_id,)).fetchall()
    return render_template("remessa_materias_primas.html", remessa=remessa, mps=mps, itens=itens)


@app.route("/produtos/<int:produto_id>/plano-corte", methods=["GET", "POST"])
def produto_plano_corte(produto_id):
    db = get_db()
    produto = db.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()
    if not produto:
        flash("Produto não encontrado.", "erro")
        return redirect(url_for("cadastros_produtos"))
    if request.method == "POST":
        def _forn_id(n):
            v = request.form.get(f"fornecedor{n}_id", "").strip()
            return int(v) if v else None
        campos = {
            "tipo_tecido":        request.form.get("tipo_tecido", "").strip() or None,
            "fornecedor1_id":     _forn_id(1),
            "fornecedor2_id":     _forn_id(2),
            "fornecedor3_id":     _forn_id(3),
            "fornecedor4_id":     _forn_id(4),
            "fornecedor5_id":     _forn_id(5),
            "marca":              request.form.get("marca", "").strip() or None,
            "colecao":            request.form.get("colecao", "").strip() or None,
            "tamanho_final":      request.form.get("tamanho_final", "").strip() or None,
            "largura_tecido":     request.form.get("largura_tecido") or None,
            "comprimento_enfesto":request.form.get("comprimento_enfesto") or None,
            "largura_corte_cm":   request.form.get("largura_corte_cm") or None,
            "comprimento_corte_cm":request.form.get("comprimento_corte_cm") or None,
            "tecido_dobrado":     1 if request.form.get("tecido_dobrado") == "1" else 0,
            "tipo_ziper":         request.form.get("tipo_ziper", "").strip() or None,
            "puxadas_min":        request.form.get("puxadas_min") or None,
            "puxadas_max":        request.form.get("puxadas_max") or None,
            "formato_produto":    request.form.get("formato_produto", "retangular"),
            "diametro_cm":        request.form.get("diametro_cm") or None,
            "qtd_redondas_por_retangular": request.form.get("qtd_redondas_por_retangular") or 1,
            "observacoes":        request.form.get("observacoes", "").strip() or None,
            "tem_frente_fundo":   1 if request.form.get("tem_frente_fundo") == "1" else 0,
            "largura_frente_cm":  request.form.get("largura_frente_cm") or None,
            "comprimento_frente_cm": request.form.get("comprimento_frente_cm") or None,
            "largura_fundo_cm":   request.form.get("largura_fundo_cm") or None,
            "comprimento_fundo_cm":  request.form.get("comprimento_fundo_cm") or None,
            "largura_fundo_menor_cm":  request.form.get("largura_fundo_menor_cm") or None,
            "comprimento_fundo_menor_cm": request.form.get("comprimento_fundo_menor_cm") or None,
        }
        plano_atual = db.execute("SELECT foto_produto FROM planos_corte WHERE produto_id = ?", (produto_id,)).fetchone()
        arquivo = request.files.get("foto_produto")
        if arquivo and arquivo.filename:
            ext = Path(arquivo.filename).suffix.lower()
            if ext in EXTENSOES_FOTO:
                nome = f"plano_{produto_id}_foto_produto{ext}"
                arquivo.save(FOTOS_PLANO_DIR / nome)
                campos["foto_produto"] = nome
            else:
                campos["foto_produto"] = plano_atual["foto_produto"] if plano_atual else None
        else:
            campos["foto_produto"] = plano_atual["foto_produto"] if plano_atual else None

        existente = db.execute("SELECT id FROM planos_corte WHERE produto_id = ?", (produto_id,)).fetchone()
        if existente:
            db.execute("""
                UPDATE planos_corte SET tipo_tecido=?,
                fornecedor1_id=?, fornecedor2_id=?, fornecedor3_id=?, fornecedor4_id=?, fornecedor5_id=?,
                marca=?, colecao=?, tamanho_final=?, largura_tecido=?, comprimento_enfesto=?,
                largura_corte_cm=?, comprimento_corte_cm=?, tecido_dobrado=?,
                tipo_ziper=?, puxadas_min=?, puxadas_max=?, formato_produto=?, diametro_cm=?,
                qtd_redondas_por_retangular=?, observacoes=?, foto_produto=?,
                tem_frente_fundo=?, largura_frente_cm=?, comprimento_frente_cm=?,
                largura_fundo_cm=?, comprimento_fundo_cm=?,
                largura_fundo_menor_cm=?, comprimento_fundo_menor_cm=?
                WHERE produto_id=?
            """, (campos["tipo_tecido"],
                  campos["fornecedor1_id"], campos["fornecedor2_id"], campos["fornecedor3_id"],
                  campos["fornecedor4_id"], campos["fornecedor5_id"],
                  campos["marca"], campos["colecao"], campos["tamanho_final"],
                  campos["largura_tecido"], campos["comprimento_enfesto"],
                  campos["largura_corte_cm"], campos["comprimento_corte_cm"], campos["tecido_dobrado"],
                  campos["tipo_ziper"], campos["puxadas_min"], campos["puxadas_max"],
                  campos["formato_produto"], campos["diametro_cm"],
                  campos["qtd_redondas_por_retangular"],
                  campos["observacoes"], campos["foto_produto"],
                  campos["tem_frente_fundo"], campos["largura_frente_cm"], campos["comprimento_frente_cm"],
                  campos["largura_fundo_cm"], campos["comprimento_fundo_cm"],
                  campos["largura_fundo_menor_cm"], campos["comprimento_fundo_menor_cm"], produto_id))
        else:
            db.execute("""
                INSERT INTO planos_corte (produto_id, tipo_tecido,
                fornecedor1_id, fornecedor2_id, fornecedor3_id, fornecedor4_id, fornecedor5_id,
                marca, colecao, tamanho_final, largura_tecido, comprimento_enfesto,
                largura_corte_cm, comprimento_corte_cm, tecido_dobrado,
                tipo_ziper, puxadas_min, puxadas_max, formato_produto, diametro_cm,
                qtd_redondas_por_retangular, observacoes, foto_produto,
                tem_frente_fundo, largura_frente_cm, comprimento_frente_cm,
                largura_fundo_cm, comprimento_fundo_cm,
                largura_fundo_menor_cm, comprimento_fundo_menor_cm)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (produto_id, campos["tipo_tecido"],
                  campos["fornecedor1_id"], campos["fornecedor2_id"], campos["fornecedor3_id"],
                  campos["fornecedor4_id"], campos["fornecedor5_id"],
                  campos["marca"], campos["colecao"], campos["tamanho_final"],
                  campos["largura_tecido"], campos["comprimento_enfesto"],
                  campos["largura_corte_cm"], campos["comprimento_corte_cm"], campos["tecido_dobrado"],
                  campos["tipo_ziper"], campos["puxadas_min"], campos["puxadas_max"],
                  campos["formato_produto"], campos["diametro_cm"],
                  campos["qtd_redondas_por_retangular"],
                  campos["observacoes"], campos["foto_produto"],
                  campos["tem_frente_fundo"], campos["largura_frente_cm"], campos["comprimento_frente_cm"],
                  campos["largura_fundo_cm"], campos["comprimento_fundo_cm"],
                  campos["largura_fundo_menor_cm"], campos["comprimento_fundo_menor_cm"]))
        db.commit()
        flash("Plano de corte salvo.", "ok")
        return redirect(url_for("produto_plano_corte", produto_id=produto_id))
    plano = db.execute("""
        SELECT pc.*,
               COALESCE(f1.nome_razao_social, pc.fornecedor1) AS forn1_label,
               COALESCE(f2.nome_razao_social, pc.fornecedor2) AS forn2_label,
               COALESCE(f3.nome_razao_social, pc.fornecedor3) AS forn3_label,
               f4.nome_razao_social AS forn4_label,
               f5.nome_razao_social AS forn5_label
        FROM planos_corte pc
        LEFT JOIN fornecedores f1 ON f1.id = pc.fornecedor1_id
        LEFT JOIN fornecedores f2 ON f2.id = pc.fornecedor2_id
        LEFT JOIN fornecedores f3 ON f3.id = pc.fornecedor3_id
        LEFT JOIN fornecedores f4 ON f4.id = pc.fornecedor4_id
        LEFT JOIN fornecedores f5 ON f5.id = pc.fornecedor5_id
        WHERE pc.produto_id = ?
    """, (produto_id,)).fetchone()
    materias_primas = db.execute("SELECT codigo, descricao FROM materias_primas ORDER BY codigo").fetchall()
    fornecedores = db.execute("SELECT * FROM fornecedores ORDER BY codigo").fetchall()
    composicao = db.execute("""
        SELECT mp.codigo, mp.descricao, pc.quantidade
        FROM produto_composicao pc
        JOIN materias_primas mp ON mp.id = pc.materia_prima_id
        WHERE pc.produto_id = ?
        ORDER BY mp.codigo
    """, (produto_id,)).fetchall()
    cron_rows = {r["numero"]: r for r in db.execute(
        "SELECT * FROM cronometragem_corte WHERE produto_id = ? ORDER BY numero", (produto_id,)
    ).fetchall()}
    cron_medidas = []
    for n in (1, 2, 3):
        r = cron_rows.get(n)
        total_min = (r["tempo_horas"] * 60 + r["tempo_minutos"]) if r else 0
        cron_medidas.append({
            "numero": n,
            "tempo_str": _format_hhmm(total_min) if (r and (r["tempo_horas"] or r["tempo_minutos"])) else "",
            "qtd_cortada": r["qtd_cortada"] if r else "",
            "num_pessoas": r["num_pessoas"] if r else "",
            "data_medicao": r["data_medicao"] if r else "",
            "nomes_pessoas": r["nomes_pessoas"] or "" if r else "",
        })
    cron_media = _calcular_media_cronometragem(cron_medidas)
    return render_template("produto_plano_corte.html", produto=produto, plano=plano,
                           materias_primas=materias_primas, fornecedores=fornecedores,
                           composicao=composicao,
                           cron_medidas=cron_medidas, cron_media=cron_media)


# ---------------------------------------------------------------------------
# Backup do banco de dados
# ---------------------------------------------------------------------------

@app.route("/backups")
def backups():
    BACKUP_DIR.mkdir(exist_ok=True)
    arquivos = sorted(BACKUP_DIR.glob("*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    lista = [
        {
            "nome": p.name,
            "tamanho_kb": round(p.stat().st_size / 1024, 1),
            "data": datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        }
        for p in arquivos
    ]
    return render_template("backups.html", backups=lista)


@app.route("/backups/novo", methods=["POST"])
def novo_backup():
    BACKUP_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = BACKUP_DIR / f"controle_remessa_{timestamp}.db"
    shutil.copy2(DB_PATH, destino)
    flash(f"Backup criado com sucesso: {destino.name}", "sucesso")
    return redirect(url_for("backups"))


@app.route("/backups/<nome>/download")
def download_backup(nome):
    return send_from_directory(BACKUP_DIR, nome, as_attachment=True)


# ---------------------------------------------------------------------------
# Ajuda
# ---------------------------------------------------------------------------

@app.route("/ajuda")
def ajuda():
    return render_template("ajuda.html")


# ---------------------------------------------------------------------------
# Download da planilha modelo gerada dinamicamente (com dados do banco)
# ---------------------------------------------------------------------------

@app.route("/planilha-modelo/download")
def download_planilha_modelo():
    import io
    import re
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COR_CABEC  = "1A1A2E"
    COR_OPC    = "2D4059"
    COR_TITULO = "E94560"
    COR_ALT1   = "FFFFFF"
    COR_ALT2   = "F0F4FF"
    COR_VERDE  = "D6F0DC"
    FONTE_BRANCA = Font(color="FFFFFF", bold=True, size=10)
    NLINHAS = 200

    def borda_fina():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def cabec(ws, col, texto, obrigatorio=True, largura=22):
        cor = COR_CABEC if obrigatorio else COR_OPC
        c = ws.cell(row=4, column=col, value=texto)
        c.fill = PatternFill("solid", fgColor=cor)
        c.font = FONTE_BRANCA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda_fina()
        ws.column_dimensions[get_column_letter(col)].width = largura

    def titulo_aba(ws, texto, subtexto):
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 48
        c = ws.cell(row=1, column=1, value=texto)
        c.font = Font(color=COR_TITULO, bold=True, size=14)
        c.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=3, column=1, value=subtexto)
        c2.font = Font(color="444444", size=9, italic=True)
        c2.fill = PatternFill("solid", fgColor="FFF8DC")
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def zebrar(ws, linha, ncols):
        cor = COR_ALT2 if linha % 2 == 0 else COR_ALT1
        for col in range(1, ncols + 1):
            c = ws.cell(row=linha, column=col)
            c.fill = PatternFill("solid", fgColor=cor)
            c.border = borda_fina()
            c.alignment = Alignment(vertical="center")

    def preencher_verde(ws, ln, ncols):
        for col in range(1, ncols + 1):
            ws.cell(ln, col).fill = PatternFill("solid", fgColor=COR_VERDE)
            ws.cell(ln, col).border = borda_fina()
            ws.cell(ln, col).alignment = Alignment(vertical="center")

    db = get_db()
    wb = openpyxl.Workbook()

    # ── ABA 0: Catálogo de Cores
    ws0 = wb.active
    ws0.title = "0-Catalogo_Cores"
    ws0.freeze_panes = "A5"
    titulo_aba(ws0, "CATALOGO DE CORES / ESTAMPAS",
        "Descricao: nome unico da cor ou estampa (ex: AZUL ROYAL, FLORAL VERAO).   "
        "Nomes duplicados serao ignorados na importacao.   "
        "Nao altere a linha 4 (cabecalho). Preencha a partir da linha 5.")
    cabec(ws0, 1, "DESCRICAO *", largura=40)
    rows0 = db.execute("SELECT descricao FROM catalogo_cores ORDER BY descricao").fetchall()
    for i, row in enumerate(rows0):
        ln = 5 + i
        ws0.cell(ln, 1, row["descricao"])
        preencher_verde(ws0, ln, 1)
    inicio0 = 5 + len(rows0)
    for ln in range(inicio0, inicio0 + NLINHAS):
        zebrar(ws0, ln, 1)

    # ── ABA 0b: Catálogo de Serviços
    ws0b = wb.create_sheet("0b-Catalogo_Servicos")
    ws0b.freeze_panes = "A5"
    titulo_aba(ws0b, "CATALOGO DE SERVICOS",
        "Descricao: nome unico do servico (ex: COSTURA E OVERLOQUE COMPLETO).   "
        "Nomes duplicados serao ignorados na importacao.   "
        "Nao altere a linha 4 (cabecalho). Preencha a partir da linha 5.")
    cabec(ws0b, 1, "DESCRICAO *", largura=50)
    rows0b = db.execute("SELECT descricao FROM catalogo_servicos ORDER BY descricao").fetchall()
    for i, row in enumerate(rows0b):
        ln = 5 + i
        ws0b.cell(ln, 1, row["descricao"])
        preencher_verde(ws0b, ln, 1)
    inicio0b = 5 + len(rows0b)
    for ln in range(inicio0b, inicio0b + NLINHAS):
        zebrar(ws0b, ln, 1)

    # ── ABA 1: Prestadores
    ws1 = wb.create_sheet("1-Prestadores")
    ws1.freeze_panes = "A5"
    titulo_aba(ws1, "PRESTADORES DE SERVICO (Terceirizados)",
        "Codigo: unico.   Nome: nome completo.   Telefone: opcional.   "
        "Registrado: 1 = com registro, 0 = sem registro.   Nao altere a linha 4. Preencha a partir da linha 5.")
    ws1.merge_cells("A3:D3")
    cabec(ws1, 1, "CODIGO *", largura=14)
    cabec(ws1, 2, "NOME *", largura=32)
    cabec(ws1, 3, "TELEFONE", obrigatorio=False, largura=20)
    cabec(ws1, 4, "REGISTRADO *", largura=14)
    rows1 = db.execute("SELECT codigo, nome, telefone, registrado FROM terceirizados ORDER BY codigo").fetchall()
    for i, row in enumerate(rows1):
        ln = 5 + i
        ws1.cell(ln, 1, row["codigo"])
        ws1.cell(ln, 2, row["nome"])
        ws1.cell(ln, 3, row["telefone"] or "")
        ws1.cell(ln, 4, row["registrado"])
        preencher_verde(ws1, ln, 4)
    inicio1 = 5 + len(rows1)
    for ln in range(inicio1, inicio1 + NLINHAS):
        zebrar(ws1, ln, 4)

    # ── ABA 2: Produtos
    ws2 = wb.create_sheet("2-Produtos")
    ws2.freeze_panes = "A5"
    titulo_aba(ws2, "PRODUTOS",
        "Codigo: unico (ex: 47, P001).   Descricao: nome do produto.   "
        "O Codigo do Produto e referenciado nas abas de Cores/Estampas e Servicos.")
    ws2.merge_cells("A3:B3")
    cabec(ws2, 1, "CODIGO *", largura=14)
    cabec(ws2, 2, "DESCRICAO *", largura=40)
    rows2 = db.execute("SELECT codigo, descricao FROM produtos ORDER BY codigo").fetchall()
    for i, row in enumerate(rows2):
        ln = 5 + i
        ws2.cell(ln, 1, row["codigo"])
        ws2.cell(ln, 2, row["descricao"])
        preencher_verde(ws2, ln, 2)
    inicio2 = 5 + len(rows2)
    for ln in range(inicio2, inicio2 + NLINHAS):
        zebrar(ws2, ln, 2)

    # ── ABA 3: Cores / Estampas
    ws3 = wb.create_sheet("3-Cores_Estampas")
    ws3.freeze_panes = "A5"
    titulo_aba(ws3, "CORES / ESTAMPAS",
        "Codigo: unico (ex: C001).   Codigo do Produto: identico ao da aba 2-Produtos.   "
        "Descricao: nome da cor ou estampa.")
    ws3.merge_cells("A3:C3")
    cabec(ws3, 1, "CODIGO *", largura=14)
    cabec(ws3, 2, "CODIGO DO PRODUTO *", largura=22)
    cabec(ws3, 3, "DESCRICAO *", largura=36)
    rows3 = db.execute("""
        SELECT ce.codigo, p.codigo AS produto_codigo, ce.descricao
        FROM cores_estampas ce JOIN produtos p ON ce.produto_id = p.id
        ORDER BY p.codigo, ce.codigo
    """).fetchall()
    for i, row in enumerate(rows3):
        ln = 5 + i
        ws3.cell(ln, 1, row["codigo"])
        ws3.cell(ln, 2, row["produto_codigo"])
        ws3.cell(ln, 3, row["descricao"])
        preencher_verde(ws3, ln, 3)
    inicio3 = 5 + len(rows3)
    maior_ce = 0
    for r in rows3:
        m = re.search(r'(\d+)$', str(r["codigo"]))
        if m:
            maior_ce = max(maior_ce, int(m.group(1)))
    fonte_sug = Font(color="AAAAAA", italic=True, size=9)
    for i, ln in enumerate(range(inicio3, inicio3 + NLINHAS)):
        zebrar(ws3, ln, 3)
        c = ws3.cell(ln, 1, "CE-{:04d}".format(maior_ce + i + 1))
        c.font = fonte_sug
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── ABA 4: Serviços
    ws4 = wb.create_sheet("4-Servicos")
    ws4.freeze_panes = "A5"
    titulo_aba(ws4, "SERVICOS",
        "Codigo: unico.   Codigo do Produto: mesmo da aba 2-Produtos.   "
        "Descricao: nome do servico.   Valor Com Registro e Valor Sem Registro: use ponto decimal (ex: 2.50).")
    ws4.merge_cells("A3:E3")
    cabec(ws4, 1, "CODIGO *", largura=14)
    cabec(ws4, 2, "CODIGO DO PRODUTO *", largura=22)
    cabec(ws4, 3, "DESCRICAO *", largura=32)
    cabec(ws4, 4, "VALOR COM REGISTRO *", largura=20)
    cabec(ws4, 5, "VALOR SEM REGISTRO *", largura=20)
    rows4 = db.execute("""
        SELECT s.codigo, p.codigo AS produto_codigo, s.descricao, s.valor_com_registro, s.valor_sem_registro
        FROM servicos s JOIN produtos p ON s.produto_id = p.id
        ORDER BY p.codigo, s.codigo
    """).fetchall()
    for i, row in enumerate(rows4):
        ln = 5 + i
        ws4.cell(ln, 1, row["codigo"])
        ws4.cell(ln, 2, row["produto_codigo"])
        ws4.cell(ln, 3, row["descricao"])
        ws4.cell(ln, 4, row["valor_com_registro"]).number_format = '#,##0.00'
        ws4.cell(ln, 5, row["valor_sem_registro"]).number_format = '#,##0.00'
        preencher_verde(ws4, ln, 5)
    inicio4 = 5 + len(rows4)
    maior_s = 0
    for r in rows4:
        m = re.search(r'(\d+)$', str(r["codigo"]))
        if m:
            maior_s = max(maior_s, int(m.group(1)))
    fonte_sug4 = Font(color="AAAAAA", italic=True, size=9)
    for i, ln in enumerate(range(inicio4, inicio4 + NLINHAS)):
        zebrar(ws4, ln, 5)
        ws4.cell(ln, 4).number_format = '#,##0.00'
        ws4.cell(ln, 5).number_format = '#,##0.00'
        c = ws4.cell(ln, 1, "S-{:04d}".format(maior_s + i + 1))
        c.font = fonte_sug4
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── ABA LEIA-ME
    ws5 = wb.create_sheet("LEIA-ME")
    ws5.sheet_properties.tabColor = "E94560"
    ws5.column_dimensions["A"].width = 90
    linhas_readme = [
        (1, "INSTRUCOES DE PREENCHIMENTO - Casa Sanchez - Importacao em Massa", True, 14, COR_TITULO),
        (3, "ORDEM DE PREENCHIMENTO", True, 11, "1A1A2E"),
        (4, "  1. Comece pela aba 2-Produtos.", False, 10, None),
        (5, "  2. Depois 3-Cores_Estampas e 4-Servicos.", False, 10, None),
        (6, "  3. Por ultimo, 1-Prestadores.", False, 10, None),
        (8, "REGRAS GERAIS", True, 11, "1A1A2E"),
        (9, "  Campos com * sao OBRIGATORIOS.", False, 10, None),
        (10, "  Codigos devem ser unicos dentro de cada aba.", False, 10, None),
        (11, "  Nao altere a linha 4 (cabecalhos azuis).", False, 10, None),
        (12, "  Linhas em VERDE ja existem no banco — inclua novas abaixo delas.", False, 10, None),
        (13, "  O sistema ignora cadastros duplicados automaticamente.", False, 10, None),
    ]
    for ln, texto, bold, size, cor in linhas_readme:
        c = ws5.cell(row=ln, column=1, value=texto)
        c.font = Font(bold=bold, size=size, color=cor or "222222")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws5.row_dimensions[ln].height = 22 if bold else 18
        if bold and cor == "1A1A2E":
            c.fill = PatternFill("solid", fgColor="E8EEFF")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="importacao_cadastros.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/aprendizado-sql")
def aprendizado_sql():
    return render_template("aprendizado_sql.html")


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
